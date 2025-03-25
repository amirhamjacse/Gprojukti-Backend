import os
import openpyxl
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from settings_management.models import OfficeLocation
from product_management.models import ProductStock, ProductPriceInfo
from user.models import UserAccount

class Command(BaseCommand):
    help = 'Import data from Excel file and create ProductStock objects'

    def handle(self, *args, **options):
        # Define your file path here
        file_path = 'settings_management/management/file/pekua_inventory.xlsx'
        self.stdout.write(f"Attempting to load Excel file from: {file_path}")

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File '{file_path}' does not exist"))
            return

        try:
            wb = load_workbook(filename=file_path)
            sheet = wb.active

            missing_products = []  # List to store missing products information

            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                # Debug print to check the row data
                # self.stdout.write(f"Processing row {row_index}: {row}")

                # if len(row) < 2:
                #     self.stdout.write(self.style.WARNING(f"Row {row_index} does not have enough columns. Skipping."))
                #     continue

                barcode = row[0]
                shop_name = row[1]  # Adjust index if your shop name is in a different column
                product_stock_qs = ProductStock.objects.filter(barcode=str(barcode)).last()
                if not product_stock_qs:
                    if barcode:
                        barcode = barcode.strip()  # Ensure whitespace is removed
                        # Find the office location based on shop name
                        office_qs = OfficeLocation.objects.filter(name__icontains=str(shop_name)).first()
                        print(office_qs)
                        
                        if not office_qs:
                            self.stdout.write(self.style.WARNING(f"No OfficeLocation found for shop name '{shop_name}'"))
                            continue
                        
                        # Assuming barcode format is like "59946-00144" and you want to split it to get product_code
                        product_code = barcode.split('-')[0]  # Adjust if your product code extraction logic is different

                        # Query ProductPriceInfo based on product code
                        product_price_info_qs = ProductPriceInfo.objects.filter(product__product_code=product_code).last()
                        print(product_price_info_qs, 'qs---------------------')

                        if not product_price_info_qs:
                            self.stdout.write(self.style.WARNING(f"No ProductPriceInfo found for product code '{product_code}'"))
                            continue
                        
                        # Find user account for 'created_by'
                        user_qs = UserAccount.objects.filter(id=1).last()
                        
                        if not user_qs:
                            self.stdout.write(self.style.WARNING(f"No UserAccount found with id 1"))
                            continue

                        # # Create ProductStock entry
                        ProductStock.objects.create(
                            barcode=barcode,
                            product_price_info=product_price_info_qs,
                            status="ACTIVE",
                            stock_location=office_qs,
                            created_by=user_qs
                        )
                        self.stdout.write(self.style.SUCCESS(f"Created ProductStock with barcode '{barcode}'"))

                    else:
                        missing_products.append({'Row': row_index, 'Barcode': barcode, 'Shop Name': shop_name})
                else:
                    office_qs = OfficeLocation.objects.filter(name__icontains=str(shop_name)).first()
                    print(office_qs, 'office_name---------')
                    product_stock_qs.status = "ACTIVE"
                    product_stock_qs.stock_location = office_qs
                    product_stock_qs.save()
                    print('Product Stock Available')

            # Handle missing products
            if missing_products:
                missing_file_path = 'missing_products.xlsx'
                wb_new = openpyxl.Workbook()
                sheet_new = wb_new.active
                sheet_new.append(['Row', 'Barcode', 'Shop Name'])

                for product in missing_products:
                    sheet_new.append([product['Row'], product['Barcode'], product['Shop Name']])

                wb_new.save(missing_file_path)
                self.stdout.write(self.style.SUCCESS(f"Created Excel file with missing products: {missing_file_path}"))

            self.stdout.write(self.style.SUCCESS("Successfully imported data from Excel"))

        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Error importing data from Excel: {e}"))
