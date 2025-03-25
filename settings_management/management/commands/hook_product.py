import openpyxl
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from settings_management.models import ShopPanelHook, ShopPanel, Product, HookProduct
from user.models import UserAccount
from utils.generates import unique_slug_generator
import os

class Command(BaseCommand):
    help = 'Import data from Excel file and create ShopPanelHook objects'

    def handle(self, *args, **options):
        ############## Must change this slug for required file #####################
        file_path = 'settings_management/management/file/Squre_self1.xlsx'  # Define your file path here
        print(f"Attempting to load Excel file from: {file_path}")

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File '{file_path}' does not exist"))
            return

        try:
            wb = load_workbook(filename=file_path)
            sheet = wb.active
            ############## Must change this slug for required panel #####################
            panel_slug = 'lifestyle-czvp-07-6976-145309-10-6976-2024'
            shop_panel = ShopPanel.objects.filter(
                slug=panel_slug).last()

            missing_products = []  # List to store missing products information

            for row in sheet.iter_rows(values_only=True):
                hook_name = row[0]
                hook_serial = row[2]
                product_slugs = row[1]

                if product_slugs:
                    product_slugs = [slug.strip() for slug in product_slugs.split(',') if slug.strip()]
                else:
                    product_slugs = ['']

                for single_slug in product_slugs:
                    if single_slug:
                        product_ins = Product.objects.filter(slug=single_slug).first()
                    else:
                        product_ins = None

                    shop_panel_hook = ShopPanelHook.objects.filter(name=hook_name, shop_panel=shop_panel).first()

                    if not shop_panel_hook:
                        shophook = ShopPanelHook.objects.create(
                            name=hook_name,
                            shop_panel=shop_panel,
                            slug=unique_slug_generator(name=hook_name),
                            created_by=UserAccount.objects.get(email='amir.hamja@gprojukti.com'),
                            serial_no=hook_serial
                        )
                    else:
                        shophook = shop_panel_hook

                    if product_ins:
                        HookProduct.objects.create(
                            product=product_ins,
                            serial_no=hook_serial,
                            created_by=UserAccount.objects.get(email='amir.hamja@gprojukti.com'),
                            slug=unique_slug_generator(name=product_ins.name),
                            shop_panel_hook=shophook
                        )
                    else:
                        missing_products.append({
                            'Hook Name': hook_name,
                            'Product Slug': single_slug,
                            'Hook Serial': hook_serial
                        })

            # Write missing products to a new Excel file
            if missing_products:
                missing_file_path = 'missing_products_life_style_panel.xlsx'
                wb_new = openpyxl.Workbook()
                sheet_new = wb_new.active
                sheet_new.append(['Hook Name', 'Product Slug', 'Hook Serial'])

                for product in missing_products:
                    sheet_new.append([product['Hook Name'], product['Product Slug'], product['Hook Serial']])

                wb_new.save(missing_file_path)
                self.stdout.write(self.style.SUCCESS(f"Created Excel file with missing products: {missing_file_path}"))

            self.stdout.write(self.style.SUCCESS("Successfully imported data from Excel"))

        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Error importing data from Excel: {e}"))
