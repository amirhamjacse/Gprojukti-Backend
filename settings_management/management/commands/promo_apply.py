import openpyxl
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from product_management.models import ProductPriceInfo, Product
from discount.models import PromoCode
from user.models import UserAccount
from utils.generates import unique_slug_generator
import os


class Command(BaseCommand):
    help = 'Import data from Excel file and create ProductPriceInfo objects'


    def handle(self, *args, **options):
        file_path = file_path = 'settings_management/management/file/promo.xlsx'
        print(f"Attempting to load Excel file from: {file_path}")

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File '{file_path}' does not exist"))
            return

        try:
            wb = load_workbook(filename=file_path)
            sheet = wb.active

            missing_products = []

            for row in sheet.iter_rows(values_only=True):
                product_code = row[3] if row[3] is not None else None
                try:
                    product_code = int(product_code)  # Attempt to convert to int
                except (ValueError, TypeError):
                    # Handle the case where product_code is not a valid integer
                    self.stderr.write(f"Skipping row: Invalid product_code '{product_code}'")
                    continue

                product = Product.objects.filter(product_code=product_code).last()
                # print(product_code, type(product_code), 'Product code -----------------')
                promo_code_name = row[9]
                if promo_code_name:
                    promo_code_name = "D" + promo_code_name
                    promo_code_names = str(promo_code_name).strip()
                    # print(promo_code_name, type(promo_code_name))

                    # print(promo_code_names, 'pr')
                    promo_code_ins = PromoCode.objects.filter(promo_code=promo_code_names).last() if promo_code_names else None
                    # print('promo-------------', promo_code_ins,)

                if product:
                    prod_price_ins = ProductPriceInfo.objects.filter(
                        product=product, product_price_type="ECOMMERCE"
                    ).first()
                    print(prod_price_ins.id, 'prod price ins')
                    prod_price_ins.promo_code = None
                    prod_price_ins.save()

                else:
                    missing_products.append({
                        'Product Code': product_code,
                        'Promo Code Name': promo_code_name
                    })

            # Handle missing products as needed
            if missing_products:
                missing_file_path = 'settings_management/management/file/missing_products.xlsx'
                wb_new = openpyxl.Workbook()
                sheet_new = wb_new.active
                sheet_new.append(['Product Code', 'Promo Code Name'])

                for product in missing_products:
                    sheet_new.append([product['Product Code'], product['Promo Code Name']])

                wb_new.save(missing_file_path)
                self.stdout.write(self.style.SUCCESS(f"Created Excel file with missing products: {missing_file_path}"))

            self.stdout.write(self.style.SUCCESS("Successfully imported data from Excel"))

        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Error importing data from Excel: {e}"))