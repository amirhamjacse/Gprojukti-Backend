from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from order.models import Order, OrderItem
import openpyxl
import os
from datetime import datetime

class Command(BaseCommand):
    help = 'Import data from Excel file and update Order and OrderItem statuses'

    def handle(self, *args, **options):
        file_path = 'settings_management/management/file/order_invoice_number.xlsx'
        print(f"Attempting to load Excel file from: {file_path}")

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File '{file_path}' does not exist"))
            return

        try:
            wb = load_workbook(filename=file_path)
            sheet = wb.active

            missing_entries = []

            for row in sheet.iter_rows(values_only=True):
                # Ensure you're within bounds of the row tuple
                if len(row) >= 2:  # Check if there are at least 3 columns in the row
                    invoice = row[0] if row[0] is not None else None
                    status = row[1] if row[1] is not None else None
                    print(f"Invoice: {invoice}, Status: {status}")

                    if invoice and status:
                        order_ins = Order.objects.filter(invoice_no=invoice).last()
                        
                        if order_ins:
                            # Update the status of the Order
                            order_ins.status = status
                            order_ins.save()
                            print(f"Updated Order {invoice} status to {status}")

                            # Update the status of related OrderItems
                            order_items = OrderItem.objects.filter(order=order_ins)
                            for item in order_items:
                                item.status = status
                                item.save()
                                print(f"Updated OrderItem {item.id} status to {status}")
                        else:
                            # Handle missing order case
                            missing_entries.append({
                                'Invoice No': invoice,
                                'Status': status
                            })
                    else:
                        # Handle missing invoice or status case
                        missing_entries.append({
                            'Invoice No': invoice,
                            'Status': status
                        })

            # Handle missing entries as needed
            if missing_entries:
                missing_file_path = 'settings_management/management/file/missing_entries.xlsx'
                wb_new = openpyxl.Workbook()
                sheet_new = wb_new.active
                sheet_new.append(['Invoice No', 'Status'])

                for entry in missing_entries:
                    sheet_new.append([entry['Invoice No'], entry['Status']])

                wb_new.save(missing_file_path)
                self.stdout.write(self.style.SUCCESS(f"Created Excel file with missing entries: {missing_file_path}"))

            self.stdout.write(self.style.SUCCESS("Successfully imported data from Excel"))

        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Error importing data from Excel: {e}"))
