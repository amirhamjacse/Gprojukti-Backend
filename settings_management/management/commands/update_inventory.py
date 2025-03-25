import openpyxl
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from product_management.models import ProductStock
import os

class Command(BaseCommand):
    help = 'Import data from Excel file and update ProductStock records'

    def handle(self, *args, **options):
        file_path = 'settings_management/management/file/prod_remove.xlsx'  # Define your file path here
        print(f"Attempting to load Excel file from: {file_path}")

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File '{file_path}' does not exist"))
            return

        try:
            wb = load_workbook(filename=file_path)
            sheet = wb.active
            
            # File to store product stock records that were not updated
            not_updated_file_path = 'product_stocks_not_updated_update_inventory.xlsx'
            not_updated_records = []  # List to store records that were not updated

            for row in sheet.iter_rows(values_only=True):
                barcode = row[0]  # Assuming barcode is in the first column

                        # Check and update the stock_location
                        # if product_stock.stock_location_id != 23:
                        # else:
                        #     # If stock_location_id is already 23, do nothing
                        #     continue
                if barcode:
                    # Check if ProductStock with the given barcode exists
                    product_stock = ProductStock.objects.filter(barcode=str(barcode)).last()

                    if product_stock:
                        product_stock.stock_location_id = 40
                        product_stock.status = "ACTIVE"
                        product_stock.save()
                    else:
                        # Add to the list of records not found
                        not_updated_records.append({
                            'Barcode': barcode,
                            'Current Stock Location': 'Not Found'
                        })

            # Write not updated records to a new Excel file
            if not_updated_records:
                wb_new = openpyxl.Workbook()
                sheet_new = wb_new.active
                sheet_new.append(['Barcode', 'Current Stock Location'])

                for record in not_updated_records:
                    sheet_new.append([record['Barcode'], record['Current Stock Location']])

                wb_new.save(not_updated_file_path)
                self.stdout.write(self.style.SUCCESS(f"Created Excel file with not updated product stocks: {not_updated_file_path}"))

            self.stdout.write(self.style.SUCCESS("Successfully processed data from Excel"))

        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Error processing data from Excel: {e}"))

class Design():
    model = ''
    


    