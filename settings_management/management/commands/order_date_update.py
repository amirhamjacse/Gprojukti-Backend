from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from order.models import Order
import openpyxl
import os
from datetime import date
from datetime import datetime

class Command(BaseCommand):
    help = 'Import data from Excel file and update Order objects'

    def handle(self, *args, **options):
        file_path = 'settings_management/management/file/last_day.xlsx'
        print(f"Attempting to load Excel file from: {file_path}")

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File '{file_path}' does not exist"))
            return

        try:
            wb = load_workbook(filename=file_path)
            sheet = wb.active

            missing_products = []

            for row in sheet.iter_rows(values_only=True):
                # Ensure you're within bounds of the row tuple
                if len(row) >= 2:  # Check if there are at least 5 columns in the row
                    invoice = row[2] if row[2] is not None else None
                    print(invoice)
                    fixed_date = date(2024, 7, 14) # Adjust index if date is in a different column
                    print(fixed_date, 'fix date')
                else:
                    # Handle case where row doesn't have enough columns
                    invoice = None
                    fixed_date = None

                if invoice:
                    print('works')
                    order_ins = Order.objects.filter(invoice_no=invoice).last()
                    # print(order_ins, 'works date')

                    if order_ins:
                        # Assuming 'date' is a string in the format 
                        print(fixed_date, 'fixed ----------')
                        parsed_date = datetime.strptime(str(fixed_date), '%Y-%m-%d').date()
                        print(parsed_date, 'works')
                        order_ins.order_date = parsed_date
                        # Uncomment to save changes to order_ins
                        order_ins.save()
                    else:
                        # Handle missing order case
                        missing_products.append({
                            'Invoice No': invoice,
                            'Order Date': date
                        })
                else:
                    # Handle missing invoice case
                    missing_products.append({
                        'Invoice No': None,
                        'Order Date': date
                    })

            # Handle missing products as needed
            if missing_products:
                missing_file_path = 'settings_management/management/file/missing_products.xlsx'
                wb_new = openpyxl.Workbook()
                sheet_new = wb_new.active
                sheet_new.append(['Invoice No', 'Order Date'])

                for product in missing_products:
                    sheet_new.append([product['Invoice No'], product['Order Date']])

                wb_new.save(missing_file_path)
                self.stdout.write(self.style.SUCCESS(f"Created Excel file with missing products: {missing_file_path}"))

            self.stdout.write(self.style.SUCCESS("Successfully imported data from Excel"))

        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Error importing data from Excel: {e}"))
