from django.db.models import Q # type: ignore
from Script.code.product import *
from Script.code.user import *

# import re
from Script.dataset.discount import *
from Script.dataset.promo_code import PROMO_CODE_LIST
from Script.dataset.stock_transfer import STOCK_TRANSFER_LIST
from base.models import *
from base.serializers import *
from base.filters import *
from courier_management.models import Courier, CourierService
from human_resource_management.models.attendance import EmployeeAttendance, EmployeeOfficeHour
from human_resource_management.models.employee import *
from location.models import *
from location.serializers import OfficeLocationListSerializer, OfficeLocationLiteSerializer
from order.models import CustomerAddressInfoLog, DeliveryMethod, Order, OrderItem, OrderItemStatusLog, OrderItemWarrantyLog, OrderPaymentLog, ServicingOrder
from product_management.models.brand_seller import Seller
from product_management.models.category import CategoryGroup
from product_management.utils import barcode_status_log
from settings_management.models import ShopDayEnd, Slider
from user.models import *
from utils.calculate import service_order_create_or_update
from utils.dataset import *
from utils.generates import generate_invoice_no, generate_service_invoice_no, unique_slug_generator_for_product_category
from utils.response_wrapper import ResponseWrapper
from utils.custom_veinlet import CustomViewSet
from rest_framework.permissions import AllowAny # type: ignore
from utils.permissions import *
import re
from django.conf import settings # type: ignore
from django_filters.rest_framework import DjangoFilterBackend # type: ignore
from rest_framework import filters # type: ignore
from utils.decorators import log_activity

import openpyxl # type: ignore
import ast
import random

import logging
import csv

from django.db.models import F

logger = logging.getLogger(__name__)


from django.core.exceptions import ObjectDoesNotExist


not_found_image_url = settings.NOT_FOUND_IMAGE


def convert_null_to_zero(value):
    return 0.0 if value == 'NULL' else value


class BulkDataUploadViewSet(CustomViewSet):
    queryset = User.objects.all()
    lookup_field = 'pk'
    serializer_class = DataSetFileSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = (
        DjangoFilterBackend,
        filters.OrderingFilter,
    )
    filterset_class = None 
    
    def bulk_command(self, request, *args, **kwargs):
        
        # .........***......... Product Price MSP, MRP Update Start .........***.........
        
        # invalid_price_infos = ProductPriceInfo.objects.filter(msp__gt=F('mrp'))
        
        # print(f"Total = {invalid_price_infos.count()}")
        
        # for price_info in invalid_price_infos:
            
        #     print(f"Before MSP = {price_info.msp}, & MRP = {price_info.mrp}")
            
        #     previous_msp = price_info.msp 
        #     price_info.msp = price_info.mrp 
        #     price_info.mrp = previous_msp 
        #     price_info.save()
            
        #     print(f"After MSP = {price_info.msp}, & MRP = {price_info.mrp}")
        
        
        # .........***......... Product Price MSP, MRP Update END .........***.........
        
        # .........***......... Bulk Barcode Entry Start .........***.........
        
        # product_code_list = ["10157", "11064", "16635", "21838", "82404", "48511", "43523"]
        
        # # product_code_list = ["10157"]
        
        # print(f"Barcode List = {product_code_list}")
        
        # serial_no = 1
        
        # stock_location_qs = OfficeLocation.objects.filter(slug = "test-shop").last()
        
        
        # for product_code in product_code_list:
            
        #     if product_code in ["10157", "11064", "16635", "21838"]:
            
        #     # if product_code in ["10157"]:
        #         product_price_info_qs = ProductPriceInfo.objects.filter(product__product_code = product_code, product_price_type = "POINT_OF_SELL").last()
        #     else:
        #         product_price_info_qs = ProductPriceInfo.objects.filter(product__product_code = product_code, product_price_type = "ECOMMERCE").last()
            
        #     for i in range(555, 609):
                
        #         if i < 10:
            
        #             barcode = f"{product_code}-0100{i}"
        #         else:
                    
        #             barcode = f"{product_code}-02{i}"
            
            
        #         print(f"Product Code = {product_code}, Barcode = {barcode}")
                
        #         product_stock_qs = ProductStock.objects.filter(barcode = barcode).last()
                
        #         if product_stock_qs:
        #             print(f"This {barcode} is Already Found in {product_stock_qs.stock_location.name}")
        #         else:
        #             print(f"This {barcode} is Creating................")
                    
        #             qs = ProductStock.objects.create(
        #                 barcode = barcode,
        #                 product_price_info = product_price_info_qs,
        #                 status = "ACTIVE",
        #                 stock_location = stock_location_qs,
        #                 stock_in_date = settings.TODAY,
        #                 created_by = request.user
        #             )
        
        # company_qs = Company.objects.all().first()
        # office_qs = OfficeLocation.objects.filter(is_shown_in_website = True, is_active = True)
        
        # for office in office_qs:
        #     slug = unique_slug_generator(name = office.name)
        #     user_qs = UserAccount.objects.filter(email__icontains = "admin@gmail.com").last()
            
        #     qs = CourierService.objects.create(name = office.name, slug = slug, courier_type = "IN_HOUSE" , company = company_qs, created_by = user_qs)
            
        today = timezone.now().date()
        
        # qs = OrderPaymentLog.objects.exclude().filter(created_at__date = today)
        
        # print(f"Total = {qs.count()}")
        
        # total = qs.count()
        # count = 1
        
        # for payment in qs:
        #     rest_of = total - count
        #     # print(f"Before Order = {payment.order.invoice_no}, Order Date = {payment.order.order_date}")
        #     payment.created_at = payment.order.order_date
        #     payment.save()
            
        #     print(f"SI ={count}, Rest Of = {rest_of} After Order = {payment.order.invoice_no}, Order Date = {payment.order.order_date}")
            
        #     count +=1
            
        start_date = "2024-07-01"
        end_date = "2025-07-01"

        qs = OrderItem.objects.filter(
            status__in=[
                "PACKAGED", "READY_FOR_PICKUP", "IN_TRANSIT", "DISPATCHED",
                "SHOP_DELIVERY_IN_TRANSIT", "SHOP_RECEIVED", 
                "DELIVERED_TO_CUSTOMER", "DELIVERED"
            ],
            order__order_date__date__range = (start_date, end_date)
        )
            
        print(f"Total =  {qs.count()}, Barcode = {qs.last().barcode_number}")
        count = 1
        for order_item in qs:
            barcode_number = order_item.barcode_number
            if barcode_number:
                product_stock_qs = ProductStock.objects.filter(barcode = barcode_number).last()
                
                if product_stock_qs:
                    if product_stock_qs.status == "ACTIVE":
                        product_stock_qs.status = "SOLD"
                        product_stock_qs.save()
                    print(f"....###....{barcode_number}, Number Update....###....")
                else:
                    print(f"....****....{barcode_number}, Number is Not Found....****....")
        
        # ProductStock.objects.filter(
        #     barcode__in=qs.values_list('barcode_number', flat=True)
        # ).update(status="SOLD")
         
        # for office in office_qs:
        #     slug = unique_slug_generator(name = office.name)
        #     user_qs = UserAccount.objects.filter(email__icontains = "admin@gmail.com").last()
            
        #     qs = CourierService.objects.create(name = office.name, slug = slug, courier_type = "IN_HOUSE" , company = company_qs, created_by = user_qs)
        
        # .........***......... Bulk Barcode Entry, MRP Update END .........***.........
        
        return ResponseWrapper(msg='Success', status=200)
    
    @log_activity
    def bulk_pre_data_create(self, request,  *args, **kwargs):
       
        subscription_qs = Subscription.objects.all()
        
        if not subscription_qs:
            subscription_qs = Subscription.objects.create(name = "Premium", slug="premium", plan='PREMIUM', plan_type='LIFETIME', created_by = request.user)
            
        if subscription_qs:
            subscription_qs = Subscription.objects.all()
            subscription_qs = subscription_qs.update(name = "Premium", slug="premium", plan='PREMIUM', plan_type='LIFETIME', created_by = request.user)
            
       
        company_type_qs = CompanyType.objects.all()
        if not company_type_qs:
            company_type_qs = CompanyType.objects.create(name = "Marketplace", slug="marketplace", created_by = request.user)
        if company_type_qs:
            company_type_qs = CompanyType.objects.all()
            company_type_qs = company_type_qs.update(name = "Marketplace", slug="marketplace", created_by = request.user)
       
        country_qs = Country.objects.all()
        if not country_qs:
            country_qs = Country.objects.create(name = "Bangladesh", slug="bangladesh", bn_name='বাংলাদেশ', created_by = request.user)
        if country_qs:
            country_qs = Country.objects.all()
            country_qs = country_qs.update(name = "Bangladesh", slug="bangladesh", bn_name='বাংলাদেশ', created_by = request.user)
       
        company_qs = Company.objects.all()
        subscription_qs = Subscription.objects.all().last()
        country_qs = Country.objects.all().last()
        company_type_qs = CompanyType.objects.all().last()
        
        if not company_qs:
            company_qs = Company.objects.create(name = "GProjukti.com", slug="gprojukti", logo='https://gprmain.sgp1.cdn.digitaloceanspaces.com/dev-gprojukti/test/8292bf8228d947f29d6ef7a6b55de87e-CategoryImage.jpg', primary_phone ='+8809612483247', secondary_phone='+8809612483247', email='info@gprojukti.com', website_url='https://www.gprojukti.com/', vat_registration_no='09876543232', registration_number= '76543235676543',address='House No: 156, Road-12, Block E, level 9,Kemal Ataturk Avenue, Banani, Dhaka -1212, Bangladesh', starting_date='2020-01-01',  subscription= subscription_qs, company_type=company_type_qs,currency='BDT',status='ACTIVE', subscription_ends='2099-12-31', created_by = request.user)
        if company_qs:
            company_qs = Company.objects.all()
            
            company_qs = company_qs.update(name = "GProjukti.com", slug="gprojukti", logo='https://gprmain.sgp1.cdn.digitaloceanspaces.com/dev-gprojukti/test/8292bf8228d947f29d6ef7a6b55de87e-CategoryImage.jpg', primary_phone ='+8809612483247', secondary_phone='+8809612483247', email='info@gprojukti.com', website_url='https://www.gprojukti.com/', vat_registration_no='09876543232', registration_number= '76543235676543',address='House No: 156, Road-12, Block E, level 9,Kemal Ataturk Avenue, Banani, Dhaka -1212, Bangladesh', starting_date='2020-01-01',  subscription= subscription_qs, company_type=company_type_qs,currency='BDT',status='ACTIVE', subscription_ends='2099-12-31', created_by = request.user)
            
        payment_type_qs = PaymentType.objects.all()
        
        for payment_type in payment_type_qs:
            company_qs = Company.objects.all().last()
            company_qs.payment_type.add(payment_type)
            
        print('Company Type Create Done') 
        
        employee_office_hour_list = EMPLOYEE_OFFICE_HOURS_LIST
        
        employee_qs = EmployeeInformation.objects.all() 
        
        for office_hour in employee_office_hour_list:
            day = office_hour.get('day')
            type = office_hour.get('type')
            start_time = office_hour.get('start_time')
            end_time = office_hour.get('end_time')
            grace_time = office_hour.get('grace_time')
            
            office_hour_qs = EmployeeOfficeHour.objects.filter(
                day= day, type=type, start_time = start_time, end_time = end_time
            )
            created_by_qs = UserAccount.objects.filter(
                id= 1
            ).last()
            
            name = f"{day}-{type}"
            slug = unique_slug_generator(name = name)
            
            if office_hour_qs:
                office_hour_qs = office_hour_qs.update(
                    day=day,
                    slug=slug,
                    type=type,
                    start_time=start_time,
                    end_time=end_time,
                    grace_time=grace_time,
                    created_by=created_by_qs,
                    )
                
            else:
                office_hour_qs = EmployeeOfficeHour.objects.create(
                    day=day,
                    slug=slug,
                    type=type,
                    start_time=start_time,
                    end_time=end_time,
                    grace_time=grace_time,
                    created_by=created_by_qs,
                    )
                
        #     employee_qs = EmployeeInformation.objects.all()
            
        #     employee_office_hour_qs = EmployeeAttendance.objects.all()
            
        #     # for employee in employee_qs:
        #     #     office_hour_qs = EmployeeOfficeHour.objects.filter(
        #     #             day= day, type=type
        #     #         ).last()
        #     #     office_hour_qs.employee_information.add(employee)
        #     #     print(f"Office Hour = {employee.name}, Hour = {office_hour_qs}")
            
        #     for employee in employee_office_hour_qs:
        #         description = """  1) Total Order Confirmed: 25 <br> 2) Cancel Order: 04 <br>
        #                 3) Outbound Call: 35 <br>
        #                 4) Inbound Call: 04 <br>
        #                 5) Pending order: 07 <br>
        #                 6) Complain issue Solve: 01 <br>
        #                 7) Facebook messages reply: 291 <br>
        #                 8) Facebook comments reply: 33 <br>
        #                 9) Outbound call for customer query: 06 <br>
        #                 10) Outbound call for Pending Order Issue: 09 <br>
        #                 11) Outbound call for order confirmation issue: 13 <br>
        #                 12) Outbound call for Order cancel Issue: 03 <br>
        #                 13) Outbound call for complain Issue: 02 <br>
        #                 14) Outbound call for Redx issue: 02 <br>
        #                 15) Contact With Redx delivery man for product Delivery issue. <br>
        #                 16) Contact With Shyamoli, malibagh & Kalshi shop for order  <brconfirmation & cancelation issue. <br>
        #                 17) Contact with Soumik bhai about product Stock Availability issue <br>
        #                 18) Contact with Taherpur shop for order Advance payment issue. <br>
        #                 19) Contact with MIS team for order correction & website issue <br>
        #                 20) Attending Ecommerce meeting. <br>
        #                 21) Discussed with Haider sir about express delivery issue."""

                
        #         qs = EmployeeAttendance.objects.filter(
        #                 id = employee.id
        #             ).last()
                      
                
        #         print('ggggggggggg', employee, qs)  
                
        #         name='fffff'
                
        #         if qs.employee_information:
        #             name = f"{qs.employee_information.name}"
                
        #         slug = unique_slug_generator(name = name)
        #         # qs = EmployeeAttendance.objects.filter(
        #         #         day= day, type=type
        #         #     ).last()
        #         qs.slug = slug
        #         qs.working_description = description
        #         qs.save()
                
        #         # office_hour_qs.employee_information.add(employee)
                
        #         print(f"Office Hour = {employee}, Hour = {slug}")

                
        return ResponseWrapper(msg='Success', status=200)
 
    @log_activity
    def bulk_payment_type_create(self, request,  *args, **kwargs):
        payment_type_list = PAYMENT_TYPE_LIST
        
        for item in payment_type_list:
            
            image_url = settings.NOT_FOUND_IMAGE
            
            name = item.get('name')
            slug = item.get('slug')
            image_url = item.get('logo')
            
            qs = PaymentType.objects.filter(slug = slug).last()
            
            if qs:
                qs.name = name
                qs.slug = slug
                qs.logo = image_url
                qs.save()
            else:
                qs = PaymentType.objects.create(
                    name=name,
                    slug = slug,
                    logo = image_url,
                    created_by = request.user
                    )
        print('Payment Type Create Done')
                
        return ResponseWrapper(msg='Success', status=200)

    @log_activity
    def bulk_delivery_method_create(self, request,  *args, **kwargs):
        delivery_method_list = DELIVERY_METHOD_LIST
        
        for item in delivery_method_list:
        
            delivery_type = item.get('delivery_type')
            slug = item.get('slug')
            delivery_charge = item.get('delivery_charge')
            
            qs = DeliveryMethod.objects.filter(slug = slug)
            
            if qs:
                qs = qs.update(
                    delivery_type=delivery_type,
                    slug = slug,
                    delivery_charge = delivery_charge,
                    created_by = request.user
                    )
            else:
                qs = DeliveryMethod.objects.create(
                    delivery_type=delivery_type,
                    slug = slug,
                    delivery_charge = delivery_charge,
                    created_by = request.user
                    )
                
        print('Create Done')
                
        return ResponseWrapper(msg='Success', status=200)

    @log_activity
    def bulk_pos_area_create(self, request,  *args, **kwargs):
        pos_area_list = POS_AREA_LIST
        
        for item in pos_area_list:
            name = item.get('name')
            slug = item.get('slug')
            bn_name = item.get('bn_name')
            district = item.get('district')
            phone = item.get('phone')
            
            print('Name = ', name)
            
            pos_area_qs = POSArea.objects.filter(name = name).last()
            
            if pos_area_qs:
                pos_area_qs.name = name
                pos_area_qs.slug = slug
                pos_area_qs.bn_name = bn_name
                pos_area_qs.phone = phone
                pos_area_qs.created_by = request.user
                pos_area_qs.save()
            else:
                pos_area_qs = POSArea.objects.create(
                    name=name,
                    slug = slug,
                    bn_name = bn_name,
                    phone = phone,
                    created_by = request.user
                    )
                
            f_qs = District.objects.filter(name = district).last()
            
            if f_qs:
                pos_area_qs.district = f_qs
                pos_area_qs.save()
                
            
            print(f"POS Phone Number = {pos_area_qs.phone}")
                
        return ResponseWrapper(msg='Success', status=200)

    @log_activity
    def bulk_pos_region_create(self, request,  *args, **kwargs):
        pos_region_list = POS_REGION_LIST
        
        for item in pos_region_list:
            name = item.get('name')
            slug = item.get('slug')
            bn_name = item.get('bn_name')
            pos_area_list = item.get('pos_area')
            
            qs = POSRegion.objects.filter(name = name).last()
            print('area', name)
            
            if qs:
                qs.name = name
                qs.slug = slug
                qs.bn_name = bn_name
                qs.created_by = request.user
                qs.save()
            else:
                qs = POSRegion.objects.create(
                    name=name,
                    slug = slug,
                    bn_name = bn_name,
                    created_by = request.user
                    )
                
            for area in pos_area_list:
                f_qs = POSArea.objects.filter(name = area.get('name')).last()
                
                
                if f_qs:
                    qs.pos_area.add(f_qs)
                
        print('POS Region Create Done')
                
        return ResponseWrapper(msg='Success', status=200)

    @log_activity
    def bulk_slider_create(self, request,  *args, **kwargs):
        slider_list = SLIDER_LIST
        
        for item in slider_list:
            
            image_url = settings.NOT_FOUND_IMAGE
            
            name = item.get('name')
            slug = item.get('slug')
            image = item.get('image')
            is_active = item.get('is_active')
            is_slider = item.get('is_slider')
            is_popup = item.get('is_popup')
            serial_no = item.get('serial_no')
            
            qs = Slider.objects.filter(slug = slug)
            
            if qs:
                qs = qs.update(
                    name=name,
                    slug = slug,
                    image = image,
                    is_active = is_active,
                    is_slider = is_slider,
                    is_popup = is_popup,
                    serial_no = serial_no,
                    created_by = request.user
                )
            else:
                qs = Slider.objects.create(
                    name=name,
                    slug = slug,
                    image = image,
                    is_active = is_active,
                    is_slider = is_slider,
                    is_popup = is_popup,
                    serial_no = serial_no,
                    created_by = request.user
                    )
                
        print('Slider Create Done')
                
        return ResponseWrapper(msg='Success', status=200)

    @log_activity
    def user_create(self, request,  *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        phone = '-'
        email = '-'
        
        # Set the batch size for pagination
        batch_size = 500
        
        # Get the total number of rows in the sheet
        total_rows = sheet.max_row
        
        # Calculate the total number of batches
        total_batches = (total_rows // batch_size) + (1 if total_rows % batch_size != 0 else 0)
        
        for batch_number in range(total_batches):
            # Calculate the start and end row for the current batch
            start_row = batch_number * batch_size + 2
            end_row = min((batch_number + 1) * batch_size, total_rows)
            
            count = 1
            
            # Iterate over rows in the current batch
            # for row_number, row in enumerate(sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True), start=start_row):
            for row_number, row in enumerate(sheet.iter_rows(min_row=12420, values_only=True), start=1):
                # Your existing code for processing each row goes here
                
                # For example:
                index = row[0]
                first_name = row[1]
                last_name = row[2]
                contact_number = row[3]
                email = row[4]
                username = row[5]
                password = row[6]
                profile_pic_url = row[7]
                gender = row[8]
                user_status = row[9]
                date_joined = row[10]
                is_active = row[11]
                is_staff = row[12]
                is_superuser = row[13]
                user_type = row[14]
                
                
                if not index:
                    break
            
                elif "@gprojukti.com" in username:
                    print(f"Pass Email Address = {username}")
                    pass
            
                else:
                    image_url = settings.NOT_FOUND_IMAGE
                
                    print(f'User Name = {username}')
                    
                    user_qs = UserAccount.objects.filter(
                        Q(email=username)
                        | Q(phone=username)
                        )
                    
                    if username.isdigit():
                        phone = username
                        
                        if email:
                            email = email
                            
                        if email == '-':
                            email = phone
                        else:   
                            email = phone
                        
                    elif type(username)== str:
                        email = username
                        
                        if contact_number:
                            phone = contact_number
                            
                        if phone == '-':
                            phone = email
                            
                        else:   
                            phone = email
                            
                    
                    # print(f"Row Number = {row_number} and Username = {username}, Contact no = {phone}, Email = {email},password = {password} , is_superuser = {is_superuser}, date_joined ={date_joined}, is_active = {is_active}, is_staff = {is_staff}, user_type = {user_type} ")
                    
                    print(f"Row Number = {row_number} and Username = {username} password = {password} , is_superuser = {is_superuser}, date_joined ={date_joined}, is_active = {is_active}, is_staff = {is_staff}, user_type = {user_type} ")
                    
                    if password is None:
                        password = ''
                        
                    if not first_name:
                        first_name = phone
                        
                    if not last_name: 
                        last_name = email
                        
                    if not user_qs:
                        qs = UserAccount.objects.create(
                            password=password,
                            first_name=first_name,
                            last_name=last_name,
                            email=email,
                            phone=phone,
                            is_active=is_active,
                            date_joined=date_joined,
                            is_staff=is_staff,
                            is_superuser=is_superuser,
                        )
                        
                    else:
                        qs = user_qs.update(
                            password=password,
                            first_name=first_name,
                            last_name=last_name,
                            email=email,
                            phone=phone,
                            is_active=is_active,
                            date_joined=date_joined,
                            is_staff=is_staff,
                            is_superuser=is_superuser,
                        )
                        
                    user_qs = User.objects.filter(email=email,
                            phone=phone).last()
                    
                    user_type_qs = UserType.objects.filter(name = user_type).last()
                    
                    user_group_qs = UserGroup.objects.filter(name = user_type).last()
                    
                    if not user_group_qs:
                        user_group_qs = UserGroup.objects.create(name = user_type)
                    
                    if user_group_qs:
                        user_qs.groups.add(user_group_qs)
                    
                    if not user_type_qs:
                        slug = unique_slug_generator(name = user_type)
                        
                        user_type_qs = UserType.objects.create(name = user_type, slug = slug, created_by = request.user)
                        
                    name = f"{user_qs.first_name} {user_qs.last_name}"
                    
                    user_info_qs = UserInformation.objects.filter(
                        user_id=user_qs.id
                    )
                    if not user_info_qs:
                        slug = unique_slug_generator(name = name)
                        
                        user_info_qs = UserInformation.objects.create(
                            name=name, user_id=user_qs.id, image = image_url, created_by = request.user
                        )
                    else:
                        user_info_qs = user_info_qs.update(
                            name=name, image = image_url
                        )
                        
                    user_information_qs = UserInformation.objects.filter(
                        user_id=user_qs.id
                    ).last()
                        
                    if user_type_qs:
                        user_information_qs.user_type = user_type_qs
                        user_information_qs.save()
                    
                # employee_qs = EmployeeInformation.objects.filter(user = user_qs).last()
                # employee_id = f"GPRO-000{count}"
                # count +=1
                # company_qs = Company.objects.all().last()
                
                # today = settings.TODAY
                
                # if not employee_qs:
                #     employee_qs = EmployeeInformation.objects.create(user = user_qs, employee_id = employee_id, image = image_url, name = name, slug = slug, employee_company = company_qs, joining_date = today, date_of_birth = today, created_by = request.user) 
                
                # if employee_qs:
                #     employee_qs = EmployeeInformation.objects.filter(user = user_qs)
                    
                #     employee_qs = employee_qs.update(user = user_qs, employee_id = employee_id, image = image_url, name = name, slug = slug, employee_company = company_qs, joining_date = today, date_of_birth = today, created_by = request.user)
                    
            # transaction.commit() 
                
        return ResponseWrapper(msg='Success', status=200)
    
    @log_activity
    def bulk_division_create(self, request,  *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            si = row[0]
            name = row[1]
            bn_name = row[2]
            
            qs = Division.objects.filter(name = name).last()
            country_qs = Country.objects.filter(name = 'Bangladesh').last()
            
            if qs:
                qs.name = name
                qs.bn_name = bn_name
                qs.country = country_qs
                qs.save()
            else:
                
                slug = unique_slug_generator(name=name)
                qs = Division.objects.create(
                    name=name,
                    slug = slug,
                    bn_name = bn_name,
                    created_by = request.user
                    )
        print('Division Create Done')
                
        return ResponseWrapper(msg='Success', status=200)
    

    @log_activity
    def bulk_district_create(self, request,  *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            si = row[0]
            name = row[1]
            bn_name = row[2]
            division = row[5]
            
            
            qs = District.objects.filter(name = name).last()
            
            f_qs = Division.objects.filter(name = division).last()
            
            if qs:
                qs.name = name
                qs.bn_name = bn_name
                qs.division = f_qs
                qs.save()
            else:
                slug = unique_slug_generator(name=name)
                qs = District.objects.create(
                    name=name,
                    slug = slug,
                    bn_name = bn_name,
                    division = f_qs,
                    created_by = request.user
                    )
                
        print('District Create Done')
                
        return ResponseWrapper(msg='Success', status=200)
    
    @log_activity
    def bulk_area_create(self, request,  *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            si = row[0]
            name = row[1]
            bn_name = row[2]
            district = row[3]
            
            
            qs = Area.objects.filter(name = name).last()
            
            f_qs = District.objects.filter(name = district).last()
            
            if qs:
                qs.name = name
                qs.bn_name = bn_name
                qs.district = f_qs
                qs.save()
            else:
                slug = unique_slug_generator(name=name)
                qs = Area.objects.create(
                    name=name,
                    slug = slug,
                    bn_name = bn_name,
                    district = f_qs,
                    created_by = request.user
                    )
                
        print('Area Create Done')
                
        return ResponseWrapper(msg='Success', status=200)
    
    @log_activity
    def bulk_office_location_create(self, request,  *args, **kwargs):
        try:
            xlsx_file = request.FILES['file']
            workbook = openpyxl.load_workbook(xlsx_file, read_only=True)
            sheet = workbook.active

            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                si = row[0]
                name = row[1]
                store_type = row[2]
                slug = row[3]
                address = row[4]
                primary_phone = row[5]
                secondary_phone = row[6]
                location = row[7]
                map_link = row[8]
                opening_time = row[9]
                closing_time = row[10]
                shown_in_website = row[11]
                off_days = row[12]
                is_active = row[13]
                area = row[14]
                pos_area_slug = row[15]

                pos_area_name = '-'
                pos_region_name = '-'

                area_qs = Area.objects.filter(name__icontains=area).last()
                company_qs = Company.objects.filter().last()
                office_location_qs = OfficeLocation.objects.filter(name=name).last()

                print(f" Name = {name}, area_qs = {area_qs}")

                if area_qs:
                    if area_qs.district:
                        pos_area_name_qs = POSArea.objects.filter(district__name__icontains=area_qs.district.name).last()
                        if pos_area_name_qs:
                            pos_area_name = pos_area_name_qs.name
                            pos_region_name = pos_area_name_qs.pos_regions.last().name

                if not area_qs:
                    area_qs = Area.objects.filter(district__name__icontains=area).last()
                    
                    if area_qs.district:
                        pos_area_name_qs = POSArea.objects.filter(district__name__icontains=area_qs.district.name).last()
                        if pos_area_name_qs:
                            pos_area_name = pos_area_name_qs.name
                            pos_region_name = pos_area_name_qs.pos_regions.last().name

                store_no = f"10{row_number}"

                if office_location_qs:
                    qs = OfficeLocation.objects.filter(name=name).last()
                    qs.name = name
                    qs.slug = slug
                    qs.store_no = store_no
                    qs.address = address
                    qs.primary_phone = primary_phone
                    qs.map_link = map_link
                    qs.opening_time = opening_time
                    qs.closing_time = closing_time
                    qs.area_id = area_qs.id
                    qs.company_id = company_qs.id
                    qs.office_type = store_type.upper()
                    qs.is_shown_in_website = bool(shown_in_website)
                    qs.is_active = bool(is_active)
                    qs.pos_area_name = pos_area_name
                    qs.pos_region_name = pos_region_name
                    qs.created_by = request.user
                    qs.save()
                else:
                    qs = OfficeLocation.objects.create(
                        name=name,
                        slug=slug,
                        store_no=store_no,
                        address=address,
                        primary_phone=primary_phone,
                        map_link=map_link,
                        opening_time=opening_time,
                        closing_time=closing_time,
                        area=area_qs,
                        company=company_qs,
                        office_type=store_type.upper(),
                        is_shown_in_website=bool(shown_in_website),
                        is_active=bool(is_active),
                        pos_area_name=pos_area_name,
                        # off_days=off_days,
                        pos_region_name=pos_region_name,
                        created_by=request.user
                    )

            print('Office Create Done')

            return ResponseWrapper(msg = 'Office locations created successfully.')
            # return JsonResponse({'status': 'success', 'message': 'Office locations created successfully.'})

        except Exception as e:
            logger.exception("An error occurred while processing the Excel file.")
            return ResponseWrapper(error_msg=f'An error occurred while processing the Excel file. {str(e)}')
        

    @log_activity
    def bulk_category_group_create(self, request,  *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            si = row[0]
            name = row[1]
            input_translation = row[2]
            slug = row[3]
            is_active = row[4]
            is_featured = row[5]
            is_show_in_ecommece = row[6]
            ordering = row[7]
            meta_title = row[8]
            meta_description = row[9]
            canonical = row[10]
            banner_image = row[11]
            created_at = row[12]
            updated_at = row[12]
            
            translation = ast.literal_eval(input_translation)
            
            company_qs = Company.objects.filter().last()
            
            qs = CategoryGroup.objects.filter(name = name).last()
            
            print(f"Row = {si}, Name = {name}, translation = {translation}")
            
            if qs:
                qs.name = name
                qs.slug = slug
                qs.translation = translation
                qs.company = company_qs
                qs.is_active = is_active
                qs.is_featured = is_featured
                qs.is_show_in_ecommece = is_show_in_ecommece
                qs.ordering = ordering
                qs.meta_title = meta_title
                qs.meta_description = meta_description
                qs.canonical = canonical
                qs.banner_image = banner_image
                qs.company = company_qs
                qs.created_at = created_at
                qs.save()
            else:
                # slug = unique_slug_generator(name=name)
                qs = CategoryGroup.objects.create(
                    name=name,
                    slug = slug,
                    translation = translation,
                    company = company_qs,
                    is_active = is_active,
                    is_featured = is_featured,
                    is_show_in_ecommece = is_show_in_ecommece,
                    ordering = ordering,
                    meta_title = meta_title,
                    meta_description = meta_description,
                    canonical = canonical,
                    banner_image = banner_image,
                    created_at = created_at,
                    created_by = request.user
                    )
                
        print('Category Group Create Done')
                
        return ResponseWrapper(msg='Success', status=200)
    
    @log_activity
    def bulk_category_create(self, request,  *args, **kwargs):
        csv_file = request.FILES['file']
        decoded_file = csv_file.read().decode('utf-8').splitlines()
        reader = csv.reader(decoded_file)
        
        # Skip the header row
        headers = next(reader)
        
        count = 1
        total = 10538
        
        for row_number, row in enumerate(reader, start=1):
            si = row[0]
            name = row[1]
            # group_name = row[2]
            # input_translation = row[3]
            slug = row[2]
            logo = row[3]
            description = row[4]
            is_active = row[5]
            is_featured = row[6]
            is_show_in_ecommece = row[7]
            # ordering = row[10]
            # featured_ordering = row[11]
            # has_child = row[12]
            # meta_title = row[13]
            # meta_description = row[14]
            # canonical = row[15]
            # banner_image = row[16]
            # created_at = row[17]
            # updated_at = row[18]
            
            # translation = ast.literal_eval(input_translation)
            
            # f_qs = CategoryGroup.objects.filter(name = group_name).last()
            
            if not si:
                break
            
            qs = Category.objects.filter(slug = slug).last()
            
            print(f"Row = {si}, Name = {name}")
            
            if qs:
                qs.name = name
                qs.slug = slug
                qs.image = logo
                qs.banner_image = logo
                # qs.translation = translation
                qs.description = description
                # qs.category_group = f_qs
                qs.is_active = is_active
                qs.is_featured = is_featured
                qs.show_in_ecommerce = is_show_in_ecommece
                # qs.ordering = ordering
                # qs.meta_title = meta_title
                # qs.meta_description = meta_description
                # qs.canonical = canonical
                # qs.banner_image = banner_image
                # qs.featured_ordering = featured_ordering
                qs.status = 'PARENT'
                # qs.created_at = created_at
                # qs.updated_at = updated_at
                qs.save()
            # else:
                # qs = Category.objects.create(
                #     name=name,
                #     slug = slug,
                #     status = 'PARENT',
                #     featured_ordering = featured_ordering,
                #     translation = translation,
                #     category_group = f_qs,
                #     image = logo,
                #     description = description,
                #     is_active = is_active,
                #     is_featured = is_featured,
                #     show_in_ecommerce = is_show_in_ecommece,
                #     ordering = ordering,
                #     meta_title = meta_title,
                #     meta_description = meta_description,
                #     canonical = canonical,
                #     banner_image = banner_image,
                #     created_at = created_at,
                #     updated_at = updated_at,
                #     created_by = request.user
                #     )
                
            print(f'Category Create Done= {qs.description}, Logo = {qs.image}')
                
        return ResponseWrapper(msg='Success', status=200)
    
    @log_activity
    def bulk_sub_category_create(self, request,  *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            si = row[0]
            name = row[1]
            category_slug = row[2]
            input_translation = row[3]
            slug = row[4]
            logo = row[5]
            description = row[6]
            is_active = row[7]
            is_featured = row[8]
            is_show_in_ecommece = row[9]
            meta_title = row[10]
            meta_description = row[11]
            canonical = row[12]
            banner_image = row[13]
            created_at = row[14]
            updated_at = row[15]
            
            translation = ast.literal_eval(input_translation)
            
            if not logo:
                logo = settings.NOT_FOUND_IMAGE
            
            f_qs = Category.objects.filter(name = category_slug).last()
            
            qs = Category.objects.filter(name = name).last()
            
            print(f"Row = {si}, Name = {name}, translation = {translation}")
            
            if qs:
                qs.name = name
                qs.slug = slug
                qs.logo = logo
                qs.translation = translation
                qs.description = description
                qs.category_parent = f_qs
                qs.is_active = is_active
                qs.is_featured = is_featured
                qs.show_in_ecommerce = is_show_in_ecommece
                qs.meta_title = meta_title
                qs.meta_description = meta_description
                qs.canonical = canonical
                qs.banner_image = banner_image
                qs.status = 'CHILD'
                qs.created_at = created_at
                qs.updated_at = updated_at
                qs.save()
            else:
                qs = Category.objects.create(
                    name=name,
                    slug = slug,
                    status = 'CHILD',
                    translation = translation,
                    category_parent = f_qs,
                    image = logo,
                    description = description,
                    is_active = is_active,
                    is_featured = is_featured,
                    show_in_ecommerce = is_show_in_ecommece,
                    meta_title = meta_title,
                    meta_description = meta_description,
                    canonical = canonical,
                    banner_image = banner_image,
                    created_at = created_at,
                    updated_at = updated_at,
                    created_by = request.user
                    )
                
        print('Sub Category Create Done')
                
        return ResponseWrapper(msg='Success', status=200)
    
    @log_activity
    def bulk_brand_create(self, request,  *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            si = row[0]
            name = row[1]
            slug = row[2]
            logo = row[3]
            is_active = row[4]
            is_featured = row[5]
            meta_title = row[6]
            meta_description = row[7]
            canonical = row[8]
            created_at = row[9]
            updated_at = row[10]
            
            # translation = ast.literal_eval(input_translation)
            
            # if not logo:
            #     logo = settings.NOT_FOUND_IMAGE
            
            # f_qs = Category.objects.filter(name = category_slug).last()
            
            qs = Brand.objects.filter(name = name).last()
            
            print(f"Row = {si}, Name = {name}")
            
            if qs:
                qs.name = name
                qs.slug = slug
                qs.logo = logo
                qs.is_active = is_active
                qs.is_featured = is_featured
                qs.is_show_in_ecommece = True
                qs.meta_title = meta_title
                qs.meta_description = meta_description
                qs.canonical = canonical
                qs.created_at = created_at
                qs.updated_at = updated_at
                qs.save()
            else:
                qs = Brand.objects.create(
                    name=name,
                    slug = slug,
                    logo = logo,
                    is_active = is_active,
                    is_featured = is_featured,
                    is_show_in_ecommece = True,
                    meta_title = meta_title,
                    meta_description = meta_description,
                    canonical = canonical,
                    created_at = created_at,
                    updated_at = updated_at,
                    created_by = request.user
                    )
                
        print('Brand Create Done')
                
        return ResponseWrapper(msg='Success', status=200)
    
    
    def bulk_supplier_create(self, request,  *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            si = row[0]
            name = row[1]
            code = row[2]
            slug = row[3]
            contact_number = row[4]
            type = row[5]
            email = row[6]
            website = row[7]
            address = row[8]
            contact_person = row[9]
            is_active = row[12]
            created_at = row[13]
            updated_at = row[14]
            
            # translation = ast.literal_eval(input_translation)
            
            if type == 'local':
                type = 'BANGLADESHI'
            else:
                type ='INTERNATIONAL'
                
            # /if not logo:
            logo = settings.NOT_FOUND_IMAGE
            
            # f_qs = Category.objects.filter(name = category_slug).last()
            
            qs = Supplier.objects.filter(name = name).last()
            
            print(f"Row = {si}, Name = {name}")
            
            if qs:
                qs.name = name
                qs.slug = slug
                qs.logo = logo
                qs.is_active = is_active
                qs.type = type
                # qs.meta_title = meta_title
                # qs.meta_description = meta_description
                # qs.canonical = canonical
                qs.created_at = created_at
                qs.updated_at = updated_at
                qs.save()
            else:
                qs = Supplier.objects.create(
                    name=name,
                    slug = slug,
                    logo = logo,
                    is_active = is_active,
                    type = type,
                    created_at = created_at,
                    updated_at = updated_at,
                    created_by = request.user
                    )
                
        print('Supplier Create Done')
                
        return ResponseWrapper(msg='Success', status=200)
    
    
    def bulk_seller_create(self, request,  *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            si = row[0]
            name = row[1]
            
            code = f"GPRO-{row_number}"
            
            
            # translation = ast.literal_eval(input_translation)
            
            type = 'BANGLADESHI'
                
            # /if not logo:
            logo = settings.NOT_FOUND_IMAGE
            
            # f_qs = Category.objects.filter(name = category_slug).last()
            
            qs = Seller.objects.filter(name = name).last()
            
            print(f"Row = {si}, Name = {name}")
            
            if qs:
                qs.name = name
                qs.code = code
                qs.type = type
                qs.logo = logo
                qs.save()
            else:
                slug = unique_slug_generator(name=name)
                qs = Seller.objects.create(
                        name=name,
                        slug = slug,
                        code = code,
                        type = type,
                        logo = logo,
                        created_by = request.user
                    )
                
        print('Seller Create Done')
                
        return ResponseWrapper(msg='Success', status=200)
    
    
    def bulk_tax_category_create(self, request,  *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            si = row[0]
            name = row[1]
            slug = row[2]
            value_in_percentage = row[3]
            is_active = True
            type = row[5]
            created_at = row[6]
            updated_at = row[7]
            
            # translation = ast.literal_eval(input_translation)
            
            # /if not logo:
            logo = settings.NOT_FOUND_IMAGE
            
            if type=='buy_tax':
                type ='BUY_TAX'
            else:
                type='SELL_TAX'
            
            f_qs = Company.objects.filter().last()
            
            qs = TaxCategory.objects.filter(name = name).last()
            
            print(f"Row = {si}, Name = {name}")
            
            if qs:
                qs.name = name
                qs.value_in_percentage = value_in_percentage
                qs.type = type
                qs.logo = logo
                qs.company = f_qs
                qs.save()
            else:
                qs = TaxCategory.objects.create(
                        name=name,
                        slug = slug,
                        value_in_percentage = value_in_percentage,
                        type = type,
                        company = f_qs,
                        is_active = is_active,
                        created_by = request.user
                    )
                
        print('Tax Category Create Done')
                
        return ResponseWrapper(msg='Success', status=200)    
    
    
    def bulk_product_create(self, request,  *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        count = 1
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        # for row_number, row in enumerate(sheet.iter_rows(min_row=292, values_only=True), start=1):
            rest_of = 1- count
            
            print(f'Row Number = {count}, Total = 7518, Rest Of = {rest_of}')
            # print(f'Row Number = {count}, Total = 7518, Rest Of = {rest_of}')
            count +=1
            
            try:
                si = row[0]
                name = row[1]
                slug = row[2]
                brand_name = row[3]
                categories_list = row[4]
                sub_categories_list = row[5]
                input_translation = row[6]
                slug = row[7]
                input_images = row[8]
                sku = row[9]
                input_specifications = row[10]
                input_meta = row[11]
                online_discount = row[12]
                offline_discount = row[13]
                minimum_stock_quantity = row[14]
                is_active = row[15]
                is_featured = row[16]
                is_top_sale = row[17]
                is_upcoming = row[18]
                is_new_arrival = row[19]
                show_on_landing_page = row[20]
                show_in_ecommerce = row[21]
                on_the_go = row[22]
                is_cart_disabled = row[23]
                is_special_day = row[24]
                rating = row[25]
                input_integrity_guaranteed = row[26]
                ordering = row[27]
                sellers_name = row[28]
                banner_message = row[29]
                meta_title = row[30]
                meta_description = row[31]
                canonical = row[32]
                og_image = row[33]
                og_title = row[34]
                og_description = row[35]
                og_url = row[36]
                is_only_pos = row[37]
                notes = row[38]
                service_warranty_value = row[39]
                service_warranty_duration_type = row[40]
                gsheba_warranty_value = row[41]
                gsheba_warranty_duration_type = row[42]
                sell_warranty_value = row[43]
                sell_warranty_duration_type = row[44]
                official_warranty_value = row[45]
                official_warranty_duration_type = row[46]
                replacement_warranty_value = row[47]
                replacement_warranty_duration_type = row[48]
                product_code = row[49]
                msp = row[50]
                mrp = row[51]
                ecommerce_mrp = row[52]
                barcode_url = row[53]
                comission_enabled = row[54]
                gsheba_amount = row[55]
                barcode = row[56]
                barcode_serial_number = row[57]
                restock_quantity = row[58]
                quantity_per_shop = row[59]
                created_at = row[60]
                updated_at = row[61]
                created_by = row[62]
                updated_by = row[63]
                
                is_out_of_stock = False
                
                discount_qs = None
                
                if not si:
                    break
                    
                
                # translation = ast.literal_eval(input_translation)
                
                # /if not logo:
                logo = settings.NOT_FOUND_IMAGE
                
                company_qs = Company.objects.filter().last() 
                
                qs = Product.objects.filter(name = name).last()
                
                status = 'PARENT'
                translation = ast.literal_eval(input_translation)
                specifications = ast.literal_eval(input_specifications)
                meta = ast.literal_eval(input_meta)
                integrity_guaranteed = ast.literal_eval(input_integrity_guaranteed)
                
                images = eval(input_images)
                
                meta_image =  not_found_image_url
                
                # created_by_qs = UserAccount.objects.filter(email=created_by).last()
                # updated_by_qs = UserAccount.objects.filter(email=updated_by).last()
                
                created_by_qs = UserAccount.objects.filter(email=request.user.email).last()
                updated_by_qs = UserAccount.objects.filter(email=request.user.email).last()
                
                short_description = meta.get('short_description')
                description = meta.get('description')
                
                brand_qs = Brand.objects.filter(name = brand_name).last()
                if not brand_qs:
                    brand_qs = None
                    
                supplier_qs = Supplier.objects.filter(name = '-').last()
                
                if not supplier_qs:
                    supplier_qs = None
                    
                # selling_tax_category_qs = TaxCategory.objects.filter(name = selling_tax_category).last()
                
                if not supplier_qs:
                    supplier_qs = None
                    
                    
                print('Product Code = ', product_code, type(product_code))
                
                if product_code == '-' or product_code == '':
                    product_code = None
                    
                try:
                    if not product_code.isdigit():
                        product_code = None
                except:
                    pass
                    
                if not images:
                    images = [
                        f'{settings.NOT_FOUND_IMAGE}',
                        f'{settings.NOT_FOUND_IMAGE}',
                            ]
                    
                    print('.............Not Found Image..................')
                    
            
                if qs:   
                    qs = Product.objects.filter(name=name)
                    
                    if product_code == '-':
                        if not qs.last().product_code == None:
                            while not product_code:
                                product_code = random.randint(10000, 99999)
                                product_code_qs = Product.objects.filter(product_code=product_code).last()
                                if product_code_qs:
                                    product_code = None
                    
                    elif product_code:
                        product_code = product_code
                        # comment:            
                    else:
                        product_code = qs.last().product_code
                    
                    qs = qs.update(
                        name=name,
                        slug = slug,
                        status = status,
                        translation = translation,
                        specifications = specifications,
                        meta = meta,
                        short_description = short_description,
                        description = description,
                        minimum_stock_quantity = minimum_stock_quantity,
                        is_featured = bool(is_featured),
                        is_top_sale = bool(is_top_sale),
                        is_upcoming = bool(is_upcoming),
                        is_new_arrival = bool(is_new_arrival),
                        is_on_the_go = bool(on_the_go),
                        is_special_day = bool(is_special_day),
                        show_on_landing_page = bool(show_on_landing_page),
                        is_active = bool(is_active),
                        rating=rating,
                        integrity_guaranteed = integrity_guaranteed,
                        product_code = product_code,
                        banner_message = banner_message,
                        images = images,
                        is_out_of_stock = bool(is_out_of_stock),
                        is_cart_disabled = bool(is_cart_disabled),
                        sku = sku,
                        video_link = 'https://www.youtube.com/watch?v=M65dQ9hb2VI',
                        meta_title = meta_title,
                        meta_image = meta_image,
                        meta_description = meta_description,
                        og_title = og_title,
                        og_image = og_image,
                        og_description = og_description,
                        og_url = og_url,
                        canonical = canonical,
                        banner_image = meta_image,
                        
                        company = company_qs,
                        brand = brand_qs,
                        supplier = supplier_qs,
                        created_at=created_at,
                        updated_at=updated_at,
                        remarks=notes,
                        is_gift_product=bool(False),
                        
                        created_by = created_by_qs,
                        updated_by = updated_by_qs,
                    )
                else:
                    while not product_code:
                        product_code = random.randint(10000, 99999)
                        product_code_qs = Product.objects.filter(product_code=product_code).last()
                        if product_code_qs:
                            product_code = None
                        
                    qs = Product.objects.create(
                            name=name,
                            slug = slug,
                            status = status,
                            translation = translation,
                            specifications = specifications,
                            is_gift_product=bool(False),
                            meta = meta,
                            short_description = short_description,
                            description = description,
                            minimum_stock_quantity = minimum_stock_quantity,
                            is_featured = bool(is_featured),
                            is_top_sale = bool(is_top_sale),
                            is_upcoming = bool(is_upcoming),
                            is_new_arrival = bool(is_new_arrival),
                            is_on_the_go = bool(on_the_go),
                            is_special_day = bool(is_special_day),
                            show_on_landing_page = bool(show_on_landing_page),
                            is_active = bool(is_active),
                            rating=rating,
                            integrity_guaranteed = integrity_guaranteed,
                            product_code = product_code,
                            banner_message = banner_message,
                            images = images,
                            is_out_of_stock = bool(is_out_of_stock),
                            is_cart_disabled = bool(is_cart_disabled),
                            sku = sku,
                            video_link = 'https://www.youtube.com/watch?v=M65dQ9hb2VI',
                            meta_title = meta_title,
                            meta_image = meta_image,
                            meta_description = meta_description,
                            og_title = og_title,
                            og_image = og_image,
                            og_description = og_description,
                            og_url = og_url,
                            canonical = canonical,
                            banner_image = meta_image,
                            
                            company = company_qs,
                            brand = brand_qs,
                            supplier = supplier_qs,
                            created_at=created_at,
                            updated_at=updated_at,
                            remarks=notes,
                            
                            created_by = created_by_qs,
                            updated_by = updated_by_qs,
                        )
                
                print(f"Row = {si}, Name = {name}, product_code = {product_code}")
                
                qs = Product.objects.filter(slug= slug).last()
                
                for category in eval(categories_list):
                    category_qs = Category.objects.filter(slug = category).last()
                    
                    if category_qs:
                        qs.category.add(category_qs)
                    
                for sub_category in eval(sub_categories_list):
                    sub_category_qs = Category.objects.filter(slug = sub_category).last()
                    
                    if category_qs:
                        qs.sub_category.add(sub_category_qs)
                        
                ecommerce_price_qs = ProductPriceInfo.objects.filter(product = qs, product_price_type = 'ECOMMERCE')
                pos_price_qs = ProductPriceInfo.objects.filter(product = qs, product_price_type = 'POINT_OF_SELL')
                corporate_price_qs = ProductPriceInfo.objects.filter(product = qs, product_price_type = 'CORPORATE')
                b2b_price_qs = ProductPriceInfo.objects.filter(product = qs, product_price_type = 'B2B')
                
                price_types = {
                    'ECOMMERCE': (ecommerce_price_qs, ecommerce_mrp),
                    'POINT_OF_SELL': (pos_price_qs, mrp),
                    'CORPORATE': (corporate_price_qs, mrp),
                    'B2B': (b2b_price_qs, mrp)
                }

                for price_type, (price_qs, default_mrp) in price_types.items():
                    print(f"Price Type = {price_type}")
                    
                    if not gsheba_amount:
                        gsheba_amount = 0.0
                    if not mrp:
                        mrp = 0.0
                    if not msp:
                        msp = 0.0
                        
                    if not default_mrp:
                        default_mrp = 0.0
                        
                    # if msp==0.0:
                    #     msp = default_mrp   
                        
                    if price_qs.exists():
                        price_instance = price_qs.first()  # Retrieve the first instance from the queryset
                        price_instance.product = qs
                        price_instance.product_price_type = price_type
                        price_instance.buying_price = msp
                        price_instance.gsheba_amount = gsheba_amount
                        price_instance.msp = msp
                        price_instance.mrp = default_mrp
                        price_instance.created_by = created_by_qs
                        price_instance.is_active = True
                        price_instance.updated_by = updated_by_qs
                        price_instance.save()
                    else:
                        price_instance = ProductPriceInfo.objects.create(
                            product=qs,
                            product_price_type=price_type,
                            buying_price=msp,
                            gsheba_amount=gsheba_amount,
                            msp=msp,
                            mrp=default_mrp,
                            is_active = True,
                            created_by=created_by_qs,
                            updated_by=updated_by_qs
                        )
                    
                if online_discount:
                    online_discount = ast.literal_eval(online_discount)
                    
                    online_discount_amount = online_discount.get('online_discount_amount')
                    online_discount_method = online_discount.get('online_discount_method')
                    
                    if online_discount_amount==0.0:
                        pass
                    
                    elif online_discount_method:
                    
                        if online_discount_method =='flat':
                            discount_name = f"{online_discount_amount} Taka Off"
                        else:
                            discount_name = f"{online_discount_amount}% Off"
                            
                        image = settings.NOT_FOUND_IMAGE
                        
                        amount_type = online_discount_method.upper()
                        discount_type = 'ONLINE'
                        discount_status = 'DISCOUNT'
                        is_for_lifetime = True
                        
                        
                        discount_qs = Discount.objects.filter(name = discount_name).last()
                        
                        if discount_qs:
                            discount_qs.name = discount_name
                            discount_qs.image = image
                            discount_qs.discount_type = discount_type
                            discount_qs.amount_type = amount_type
                            discount_qs.discount_status = discount_status
                            discount_qs.is_for_lifetime = is_for_lifetime
                            discount_qs.company = company_qs
                            discount_qs.discount_amount = online_discount_amount
                            discount_qs.save()
                            
                        else:
                            slug = unique_slug_generator(name = discount_name)
                            
                            discount_qs = Discount.objects.create( 
                                name = discount_name,
                                slug=slug,
                                image = image,
                                discount_type = discount_type,
                                amount_type = amount_type,
                                discount_status = discount_status,
                                is_for_lifetime = is_for_lifetime,
                                company = company_qs,
                                created_by = created_by_qs,
                                discount_amount = online_discount_amount
                            )
                            
                            
                    ecom_price_qs = ProductPriceInfo.objects.filter(product = qs, product_price_type = 'ECOMMERCE').last() 
                    
                    if discount_qs:
                        
                        print(f'AAAAAAA')
                        discount_qs  = Discount.objects.filter(name = discount_name).last()
                        
                        if ecom_price_qs:
                            
                            print(f'BBBBBB')
                            ecom_price_qs.discount = discount_qs
                            ecom_price_qs.save()
                            
                            print(f'Discount Add = {discount_name}')
                                
                    # try:
                        
                    #     if ecom_price_qs:
                    #         ecommerce_price_qs.discount = discount_qs
                    #         ecom_price_qs.save()
                    # except:
                    #     pass
            
                gsheba_warranty_qs = ProductWarranty.objects.filter(product=qs, warranty_type='1_GSHEBA_WARRANTY')
                company_warranty_qs = ProductWarranty.objects.filter(product=qs, warranty_type='2_COMPANY_WARRANTY')
                replacement_warranty_qs = ProductWarranty.objects.filter(product=qs, warranty_type='3_REPLACEMENT_WARRANTY')
                service_warranty_qs = ProductWarranty.objects.filter(product=qs, warranty_type='4_SERVICE_WARRANTY')
                supplier_warranty_qs = ProductWarranty.objects.filter(product=qs, warranty_type='5_SUPPLIER_SERVICE_WARRANTY')
                
                warranty_types = {
                    '1_GSHEBA_WARRANTY': (gsheba_warranty_qs, gsheba_warranty_value, gsheba_warranty_duration_type),
                    '2_COMPANY_WARRANTY': (company_warranty_qs, official_warranty_value, official_warranty_duration_type),
                    '3_REPLACEMENT_WARRANTY': (replacement_warranty_qs, replacement_warranty_value, replacement_warranty_duration_type),
                    '4_SERVICE_WARRANTY': (service_warranty_qs, service_warranty_value, service_warranty_duration_type),
                    '5_SUPPLIER_SERVICE_WARRANTY': (supplier_warranty_qs, sell_warranty_value, sell_warranty_duration_type)
                }

                for warranty_type, (warranty_qs, value, duration_type) in warranty_types.items():
                    if value:
                        try:
                            if int(value) > 0:
                                if warranty_qs.exists():
                                    warranty_instance = warranty_qs.first()
                                    warranty_instance.value = value
                                    warranty_instance.warranty_duration = duration_type.upper()
                                    warranty_instance.product = qs
                                    warranty_instance.created_at = created_at
                                    warranty_instance.updated_at = updated_at
                                    warranty_instance.created_by = created_by_qs
                                    warranty_instance.updated_by = updated_by_qs
                                    warranty_instance.save()
                                else:
                                    warranty_instance = ProductWarranty.objects.create(
                                        product=qs,
                                        warranty_type=warranty_type,
                                        value=value,
                                        warranty_duration=duration_type.upper(),
                                        created_at=created_at,
                                        updated_at=updated_at,
                                        created_by=created_by_qs,
                                        updated_by=updated_by_qs
                                    )
                                    
                        except:
                            pass
                    
                    
            except:
                pass

        print('Product Create Done')
                
        return ResponseWrapper(msg='Success', status=200) 
    
    def bulk_product_stock_create(self, request, *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active
        
        count = 1
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            si = row[0]
            product_name = row[1]
            product_slug = row[2]
            product_code = row[3]
            shop_slug = row[7]
            total_quantity = row[8]
            barcode_list_str = row[9]
            
            if not si:
                break
            
            # Clean up barcode_list_str to get list of barcodes
            barcode_list = [barcode.strip().strip('"') for barcode in barcode_list_str.strip('[]').split(',')]

            # Assuming Product, OfficeLocation and ProductPriceInfo models are appropriately defined
            
            product_qs = Product.objects.filter(name=product_name).last()
            shop_qs = OfficeLocation.objects.filter(name__icontains=shop_slug).last()
            
            
            product_price_info_qs = ProductPriceInfo.objects.filter(product=product_qs, product_price_type='POINT_OF_SELL').last()
            
            try:
                if shop_qs.name == 'GProjukti.com - E commerce Shop':
                    product_price_info_qs = ProductPriceInfo.objects.filter(product=product_qs, product_price_type='ECOMMERCE').last()
                elif shop_qs.name == 'B2B':
                    product_price_info_qs = ProductPriceInfo.objects.filter(product=product_qs, product_price_type='B2B').last()
                    
            except:
                if not product_price_info_qs:
                    pass
                
                
            rest_of = 17082 - count
            
            print(f"......***......SI = {count} Rest of ={rest_of} and Shop Name = {shop_qs.name}......***......")
            count +=1

            if product_qs and shop_qs and product_price_info_qs:
                for barcode in barcode_list:
                    try:
                        product_stock_qs = ProductStock.objects.get(barcode=barcode)
                        print(f"Updating barcode: {barcode}")
                        # Update existing ProductStock entry
                        product_stock_qs.product_price_info = product_price_info_qs
                        product_stock_qs.status = 'ACTIVE'
                        product_stock_qs.stock_location = shop_qs
                        product_stock_qs.save()
                        
                    except ProductStock.DoesNotExist:
                        print(f"Creating barcode: {barcode}")
                        # Create new ProductStock entry
                        ProductStock.objects.create(
                            barcode=barcode,
                            product_price_info=product_price_info_qs,
                            status='ACTIVE',
                            stock_location=shop_qs,
                            created_by=request.user
                        )
            
        print('Product Stock Create/Update Done')
        return ResponseWrapper(msg='Success', status=200)
    
        
    def bulk_product_stock_create_for_warehouse(self, request, *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active
        
        count = 1
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1): 
            barcode = row[0]
            
            if not barcode:
                break
            
            product_code = barcode.split('-')[0]
            
            product_qs = Product.objects.filter(product_code=product_code).last()
            
            if not product_qs:
                print(f" {product_code} Product Code is not Found ")
                
            shop_qs = OfficeLocation.objects.filter(name__icontains="GProjukti.Com - Warehouse").last()
            
            
            product_price_info_qs = ProductPriceInfo.objects.filter(product=product_qs, product_price_type='POINT_OF_SELL').last()
            print(f'*****. {product_price_info_qs}')
            
                
            if product_price_info_qs:
                rest_of = 207 - count
                
                print(f"......***......SI = {count} Rest of ={207} and Shop Name = {shop_qs.name} Product Code {product_code}......***......")
                count +=1
                
                product_stock_qs = ProductStock.objects.filter(barcode=barcode).last()
                
                if product_stock_qs:

                # try:
                    print(f"Updating barcode: {barcode}")
                    
                    # Update existing ProductStock entry
                    
                    product_stock_qs = ProductStock.objects.filter(barcode=barcode).last()
                    
                    product_stock_qs.product_price_info = product_price_info_qs
                    product_stock_qs.status = 'ACTIVE'
                    product_stock_qs.stock_location = shop_qs
                    product_stock_qs.save()
                
                # except ProductStock.DoesNotExist:
                
                else:
                    print(f"Creating barcode: {barcode}")
                
                    # Create new ProductStock entry
                    ProductStock.objects.create(
                        barcode=barcode,
                        product_price_info=product_price_info_qs,
                        status='ACTIVE',
                        stock_location=shop_qs,
                        created_by=request.user
                    )
                
        print('Product Stock Create/Update Done')
        return ResponseWrapper(msg='Success', status=200)
    
    def bulk_promo_code_create(self, request,  *args, **kwargs):
        promo_code_list = PROMO_CODE_LIST
        
        company_qs = Company.objects.filter().last()
        
        for item in promo_code_list:
            promo_code = item.get('code')
            image = item.get('image')
            created_at = item.get('created_at')
            updated_at = item.get('updated_at')
            created_by = item.get('created_by').get('username')
            discount_amount = item.get('discount_amount')
            discount_type = item.get('discount_type')
            start_date = item.get('start_date')
            end_date = item.get('end_date')
            
            promo_type = ''
            
            product_variant = item.get('product_variant')
            
            if product_variant == []:
                promo_type = 'ECOMMERCE_SELL'
                
            elif not product_variant == []:
                promo_type = 'POINT_OF_SELL'
            
            if discount_type == 'flat':
                discount_type = 'FLAT'
            else:
                discount_type = 'PERCENTAGE'
                
            user_qs = UserAccount.objects.filter().last()
                
            promo_qs = PromoCode.objects.filter(promo_code = promo_code)
            
            if promo_qs:
                promo_qs = promo_qs.update(
                    promo_code = promo_code,
                    image = image,
                    start_date = start_date,
                    end_date = end_date,
                    schedule_type = 'DATE_WISE',
                    amount_type = discount_type,
                    discount_amount = discount_amount,
                    promo_type = promo_type,
                    company = company_qs,
                    created_by = user_qs,
                    created_at = created_at
                )
            
            elif not promo_qs:
                slug = unique_slug_generator(name = promo_code)
                
                promo_qs = PromoCode.objects.create(
                        promo_code = promo_code,
                        slug = slug,
                        image = image,
                        start_date = start_date,
                        end_date = end_date,
                        schedule_type = 'DATE_WISE',
                        amount_type = discount_type,
                        discount_amount = discount_amount,
                        promo_type = promo_type,
                        company = company_qs,
                        created_by = user_qs,
                        created_at = created_at
                    )
                
            
            p_qs = PromoCode.objects.filter(promo_code = promo_code).last()
            
            print(f"Promo Code = {promo_code},Type={p_qs.promo_type}, Amount = {discount_amount}")
            
            
            if not product_variant == []:
                for product in product_variant:
                    product_slug = product.get('slug')
                    
                    product_price_qs = None
                    
                    if p_qs.promo_type == 'POINT_OF_SELL':
                        product_price_qs = ProductPriceInfo.objects.filter(product_price_type = 'POINT_OF_SELL', product__slug = product_slug).last()
                        
                    elif p_qs.promo_type == 'ECOMMERCE_SELL':
                        product_price_qs = ProductPriceInfo.objects.filter(product_price_type = 'ECOMMERCE', product__slug = product_slug).last()
                    
                    if product_price_qs:
                        product_price_qs.promo_code = p_qs
                        product_price_qs.save()
                        
                        print(f"Promo Code ={product_slug}, Name = {product_price_qs.product.name}")
                    
        promo_code_qs = PromoCode.objects.all()
        
        return ResponseWrapper(msg=f"Total Promo Created = {promo_code_qs.count()}",status=200)
     
    # def bulk_product_stock_create(self, request,  *args, **kwargs):
    #     xlsx_file = request.FILES['file']
    #     workbook = openpyxl.load_workbook(xlsx_file)
    #     sheet = workbook.active 
        
    #     for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
    #         si = row[0]
    #         product_name = row[1]
    #         product_slug = row[2]
    #         shop_slug = row[3]
    #         total_quantity = row[4]
    #         barcode_list_str = row[5]
            
            
    #         product_qs = Product.objects.filter(name = product_name).last()
    #         shop_qs = OfficeLocation.objects.filter(name__icontains = shop_slug).last()
             
    #         barcode_list =  barcode_list_str.strip("[]").split(',')
            
    #         product_price_info_qs = ProductPriceInfo.objects.filter(product = product_qs, product_price_type='POINT_OF_SELL').last()
            
    #         if shop_qs:
    #             try:
    #                 if shop_qs.name == 'GProjukti.com - E commerce Shop':
    #                     product_price_info_qs = ProductPriceInfo.objects.filter(product = product_qs, product_price_type='ECOMMERCE').last()
                    
    #                 elif shop_qs.name == 'B2B':
    #                     product_price_info_qs = ProductPriceInfo.objects.filter(product = product_qs, product_price_type='B2B').last()
                
    #             except:
    #                 if not product_price_info_qs:
    #                     pass
    #                 # return ResponseWrapper(error_msg=f'For {product_name} Product Price is Not Found', error_code=400) 
                
    #             if product_price_info_qs:
    #                 for barcode_str in barcode_list:
                        
    #                     if barcode_str.startswith('"') and barcode_str.endswith('"'):
    #                         barcode = barcode_str[1:-1]
                        
    #                         print(f"SI NO = {si}, Barcode = {barcode}, Total Stock = {total_quantity}, Product Name = {product_name}")
                            
    #                         product_stock_qs = ProductStock.objects.filter(
    #                             barcode = barcode
    #                         ).last()
                            
    #                         status = 'ACTIVE'
                            
    #                         if product_stock_qs:
    #                             qs = ProductStock.objects.filter(
    #                             barcode = barcode 
    #                         )
    #                             product_stock_qs= qs.update(
    #                                 barcode = barcode, 
    #                                 product_price_info = product_price_info_qs, 
    #                                 status = status, 
    #                                 stock_location = shop_qs,
    #                                 created_by = request.user
    #                             )
                                
    #                         else:
    #                             product_stock_qs = ProductStock.objects.create(
    #                             barcode = barcode, 
    #                             product_price_info = product_price_info_qs, 
    #                             status = status, 
    #                             stock_location = shop_qs,
    #                             created_by = request.user
    #                         )
                                
    #                         product_stock_qs = ProductStock.objects.filter(
    #                             barcode = barcode
    #                         ).last()
                            
    #                         product_stock_log_qs = ProductStockLog.objects.filter(
    #                             product_stock = product_stock_qs, current_status = current_status
    #                         ).last()
                             
    #                         stock_location_info = OfficeLocationListSerializer(shop_qs).data
                            
    #                         current_status = product_stock_qs.status
    #                         current_status_display = product_stock_qs.get_status_display()
                            
    #                         previous_status = '-'
    #                         previous_status_display = '-'
                            
    #                         stock_status_change_by_info = BaseSerializer(request.user).data
                            
    #                         if product_stock_log_qs:
    #                             qs = ProductStockLog.objects.filter(
    #                                 product_stock = product_stock_qs
    #                             )
                                
    #                             product_stock_log_qs= qs.update(
    #                                 product_stock = product_stock_qs, 
    #                                 stock_location_info = stock_location_info, 
    #                                 current_status = current_status, 
    #                                 current_status_display = current_status_display,
    #                                 previous_status = previous_status,
    #                                 previous_status_display = previous_status_display,
    #                                 stock_status_change_by_info = stock_status_change_by_info,
    #                                 stock_in_date = product_stock_qs.created_at,
    #                                 created_by = request.user
    #                             )
                                
    #                         elif not product_stock_log_qs:
    #                             product_stock_log_qs = ProductStockLog.objects.create(
    #                                 product_stock = product_stock_qs, 
    #                                 stock_location_info = stock_location_info, 
    #                                 current_status = current_status, 
    #                                 current_status_display = current_status_display,
    #                                 previous_status = previous_status,
    #                                 previous_status_display = previous_status_display,
    #                                 stock_status_change_by_info = stock_status_change_by_info,
    #                                 stock_in_date = product_stock_qs.created_at,
    #                                 created_by = request.user
    #                         )
                  
    #     print('Product Stock Create Done')  
                
    #     return ResponseWrapper(msg='Success', status=200)
    
    
    def bulk_stock_transfer_create(self, request,  *args, **kwargs):
        # stock_transfer_list = STOCK_TRANSFER_LIST
        
        company_qs = Company.objects.filter().last()
        
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        count = 1
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=215, values_only=True), start=1):
            rest_of = 4022-count
            
            print(f"SI = {count}, Total = {4022}, Rest Of = {rest_of}")
            count +=1
            
            
            id = row[0]
            created_date = row[1]
            from_store_name = row[2]
            to_store_name = row[3]
            sku_list = row[4]
            approved_date = row[5]
            approval_status = row[6]
            received_status = row[7]
            received_sku_list = row[8]
            not_received_sku_list = row[9]
            excess_sku_list = row[10]
            
            # id = item.get('id')
            # created_date = item.get('date')
            # from_store_name = item.get('from_store_name').get('name')
            # to_store_name = item.get('to_store_name').get('name')
            # sku_list = item.get('sku_list')
            # approved_date = item.get('approved_date')
            # approval_status = item.get('approval_status')
            # received_status = item.get('received_status')
            # received_sku_list = item.get('received_sku_list')
            # not_received_sku_list = item.get('not_received_sku_list')
            # excess_sku_list = item.get('excess_sku_list')
            
            print(f"Id = {id}")
            
            product_stock_list = []
            created_by_qs = UserAccount.objects.filter(email = request.user).last()
            employee_qs = None
                
            stock_location_qs = OfficeLocation.objects.filter(name = from_store_name).last()
            to_stock_location_qs = OfficeLocation.objects.filter(name = to_store_name).last()
                
            employee_qs = EmployeeInformation.objects.filter(
                work_station = stock_location_qs
            ).last() 
            
            if employee_qs:
                created_by_qs = UserAccount.objects.filter(email = employee_qs.user.email).last()
                
            elif from_store_name == 'GProjukti.com - Warehouse':
                
                created_by_qs = UserAccount.objects.filter(email = 'haider.chowdhury@gprojukti.com').last()
                
            else:
                created_by_qs = UserAccount.objects.filter(email = request.user).last()
                        
                        
            if not employee_qs:
                employee_qs = EmployeeInformation.objects.filter(user = created_by_qs).last()
            
            # new_sku_list = [f"{sku // 1000000:05d}-{sku % 1000000:05d}" for sku in sku_list]
            
            # new_sku_list = [f"{sku // 100000}-{sku % 100000:05d}" for sku in sku_list]
            
            if not id:
                break
            
            new_sku_list = sku_list.strip('{}').split(',')
            
            print('new_sku_list', new_sku_list)
            
            for barcode in new_sku_list:
                try:
                    barcode = barcode          
                    product_stock_qs = ProductStock.objects.filter(barcode = barcode).last()
                    
                    if not product_stock_qs:
                        product_code = barcode.split('-')[0]
                        
                        product_price_info_qs = ProductPriceInfo.objects.filter(product__product_code = product_code, product_price_type = 'POINT_OF_SELL').last()
                        
                        if not product_price_info_qs:
                            print(f"{barcode} is Not Found")
                            pass
                            # return ResponseWrapper(msg = f"{barcode} is Not Found")
                        
                        status = 'ACTIVE'
                        status_display = 'Active'
                        
                        
                        stock_in_date = settings.TODAY
                        
                        print(f"******...BarCode Create.....********")
                        
                        product_stock_qs = ProductStock.objects.create(
                                barcode = barcode,
                                product_price_info = product_price_info_qs,
                                status = status,
                                stock_location = stock_location_qs,
                                stock_in_date = stock_in_date,
                                created_by = request.user
                            )
                        
                        product_stock_qs = ProductStock.objects.filter(
                                    barcode = barcode
                                ).last() 
                        
                        stock_location_info = OfficeLocationListSerializer(stock_location_qs).data
                        
                        current_status = product_stock_qs.status
                        current_status_display = product_stock_qs.get_status_display()
                        
                        previous_status = '-'
                        previous_status_display = '-'
                        
                        stock_status_change_by_info = BaseSerializer(created_by_qs).data
                        
                        product_stock_log_qs = ProductStockLog.objects.filter(
                                product_stock = product_stock_qs, current_status = current_status
                            ).last()
                        
                        if product_stock_log_qs:
                            qs = ProductStockLog.objects.filter(
                                product_stock = product_stock_qs,
                                current_status = current_status
                            )
                            
                            product_stock_log_qs= qs.update(
                                product_stock = product_stock_qs, 
                                stock_location_info = stock_location_info, 
                                current_status = current_status, 
                                current_status_display = current_status_display,
                                previous_status = previous_status,
                                previous_status_display = previous_status_display,
                                stock_status_change_by_info = stock_status_change_by_info,
                                stock_in_date = product_stock_qs.created_at,
                                created_by = request.user
                            )
                            
                        elif not product_stock_log_qs:
                            product_stock_log_qs = ProductStockLog.objects.create(
                                product_stock = product_stock_qs, 
                                stock_location_info = stock_location_info, 
                                current_status = current_status, 
                                current_status_display = current_status_display,
                                previous_status = previous_status,
                                previous_status_display = previous_status_display,
                                stock_status_change_by_info = stock_status_change_by_info,
                                stock_in_date = product_stock_qs.created_at,
                                created_by = request.user
                        )
                            
                    if product_stock_qs:
                        product_stock_qs = ProductStock.objects.filter(
                                    barcode = barcode
                                ).last() 
                        
                        print(f"######...BarCode Update.....#####")
                        
                        stock_location_info = OfficeLocationLiteSerializer(stock_location_qs).data
                        
                        current_status = product_stock_qs.status
                        current_status_display = product_stock_qs.get_status_display()
                        
                        previous_status = '-'
                        previous_status_display = '-'
                        
                        stock_status_change_by_info = BaseSerializer(created_by_qs).data
                        
                        product_stock_log_qs = ProductStockLog.objects.filter(
                                product_stock = product_stock_qs, current_status = current_status
                            ).last()
                        
                        # if product_stock_log_qs:
                        #     qs = ProductStockLog.objects.filter(
                        #         product_stock = product_stock_qs,
                        #         current_status = current_status
                        #     )
                            
                        #     product_stock_log_qs= qs.update(
                        #         product_stock = product_stock_qs, 
                        #         stock_location_info = stock_location_info, 
                        #         current_status = current_status, 
                        #         current_status_display = current_status_display,
                        #         previous_status = previous_status,
                        #         previous_status_display = previous_status_display,
                        #         stock_status_change_by_info = stock_status_change_by_info,
                        #         stock_in_date = product_stock_qs.created_at,
                        #         created_by = request.user
                        #     )
                            
                        # elif not product_stock_log_qs:
                        #     product_stock_log_qs = ProductStockLog.objects.create(
                        #         product_stock = product_stock_qs, 
                        #         stock_location_info = stock_location_info, 
                        #         current_status = current_status, 
                        #         current_status_display = current_status_display,
                        #         previous_status = previous_status,
                        #         previous_status_display = previous_status_display,
                        #         stock_status_change_by_info = stock_status_change_by_info,
                        #         stock_in_date = product_stock_qs.created_at,
                        #         created_by = request.user
                        # )
                        
                    print(f".................***................. This {barcode} Add Done In Inventory .................***................. ")
                    
                    if approval_status == 'False':
                        status = "IN_TRANSFER"
                        
                        previous_status = product_stock_qs.status
                        previous_status_display = product_stock_qs.get_status_display()
                        
                        product_stock_qs.status = status
                        product_stock_qs.save()
                        
                        # print(f"......***...... This {barcode} Status Is {product_stock_qs.get_status_display()} ......***...... ")
                        
                        stock_location_info = OfficeLocationListSerializer(product_stock_qs.stock_location).data
                        
                        current_status = product_stock_qs.status
                        current_status_display = product_stock_qs.get_status_display()
                        
                        stock_status_change_by_info = BaseSerializer(created_by_qs).data
                        
                        
                        product_stock_log_qs = ProductStockLog.objects.filter(
                                product_stock = product_stock_qs, current_status = current_status
                            ).last()
                        
                        # if product_stock_log_qs:
                        #     qs = ProductStockLog.objects.filter(
                        #         product_stock = product_stock_qs
                        #     )
                            
                        #     product_stock_log_qs= qs.update(
                        #         product_stock = product_stock_qs, 
                        #         stock_location_info = stock_location_info, 
                        #         current_status = current_status, 
                        #         current_status_display = current_status_display,
                        #         previous_status = previous_status,
                        #         previous_status_display = previous_status_display,
                        #         stock_status_change_by_info = stock_status_change_by_info,
                        #         stock_in_date = created_date,
                        #         created_by = request.user
                        #     )
                            
                        # elif not product_stock_log_qs:
                        #     product_stock_log_qs = ProductStockLog.objects.create(
                        #         product_stock = product_stock_qs, 
                        #         stock_location_info = stock_location_info, 
                        #         current_status = current_status, 
                        #         current_status_display = current_status_display,
                        #         previous_status = previous_status,
                        #         previous_status_display = previous_status_display,
                        #         stock_status_change_by_info = stock_status_change_by_info,
                        #         stock_in_date = created_date,
                        #         created_by = request.user
                        # )
                            
                    elif approval_status == 'True':
                        status = "IN_TRANSFER"
                        
                        previous_status = product_stock_qs.status
                        previous_status_display = product_stock_qs.get_status_display()
                        
                        product_stock_qs.status = status
                        
                        # product_stock_qs.stock_location = to_stock_location_qs
                        product_stock_qs.save()
                        
                        # print(f"......###...... This {barcode} Status Is {product_stock_qs.get_status_display()} ......###...... ")
                        
                        stock_location_info = OfficeLocationListSerializer(product_stock_qs.stock_location).data
                        
                        current_status = product_stock_qs.status
                        current_status_display = product_stock_qs.get_status_display()
                        
                        stock_status_change_by_info = BaseSerializer(created_by_qs).data
                        
                        product_stock_log_qs = ProductStockLog.objects.filter(
                                product_stock = product_stock_qs, current_status = current_status
                            ).last()
                        
                        # if product_stock_log_qs:
                        #     qs = ProductStockLog.objects.filter(
                        #         product_stock = product_stock_qs
                        #     )
                            
                        #     product_stock_log_qs= qs.update(
                        #         product_stock = product_stock_qs, 
                        #         stock_location_info = stock_location_info, 
                        #         current_status = current_status, 
                        #         current_status_display = current_status_display,
                        #         previous_status = previous_status,
                        #         previous_status_display = previous_status_display,
                        #         stock_status_change_by_info = stock_status_change_by_info,
                        #         stock_in_date = created_date,
                        #         created_by = request.user
                        #     )
                            
                        # elif not product_stock_log_qs:
                        #     product_stock_log_qs = ProductStockLog.objects.create(
                        #         product_stock = product_stock_qs, 
                        #         stock_location_info = stock_location_info, 
                        #         current_status = current_status, 
                        #         current_status_display = current_status_display,
                        #         previous_status = previous_status,
                        #         previous_status_display = previous_status_display,
                        #         stock_status_change_by_info = stock_status_change_by_info,
                        #         stock_in_date = created_date,
                        #         created_by = request.user
                        # )
                            
                        status = "ACTIVE"
                        
                        previous_status = product_stock_qs.status
                        previous_status_display = product_stock_qs.get_status_display()
                        
                        product_stock_location_qs = OfficeLocation.objects.filter(name = to_store_name).last()
                        
                        product_stock_qs.status = status
                        product_stock_qs.stock_location = product_stock_location_qs
                        product_stock_qs.save()
                        
                        # print(f"......###...... This {barcode} Status Is {product_stock_qs.get_status_display()} and Location is {product_stock_location_qs.name} ......###...... ")
                        
                        stock_location_info = OfficeLocationListSerializer(product_stock_location_qs).data
                        
                        current_status = product_stock_qs.status
                        current_status_display = product_stock_qs.get_status_display()
                        
                        stock_status_change_by_info = BaseSerializer(created_by_qs).data
                        
                        # if product_stock_log_qs:
                        #     qs = ProductStockLog.objects.filter(
                        #         product_stock = product_stock_qs, current_status = "ACTIVE"
                        #     )
                            
                        #     product_stock_log_qs= qs.update(
                        #         product_stock = product_stock_qs, 
                        #         stock_location_info = stock_location_info, 
                        #         current_status = current_status, 
                        #         current_status_display = current_status_display,
                        #         previous_status = previous_status,
                        #         previous_status_display = previous_status_display,
                        #         stock_status_change_by_info = stock_status_change_by_info,
                        #         stock_in_date = approved_date,
                        #         created_by = request.user
                        #     )
                            
                        # elif not product_stock_log_qs:
                        #     product_stock_log_qs = ProductStockLog.objects.create(
                        #         product_stock = product_stock_qs, 
                        #         stock_location_info = stock_location_info, 
                        #         current_status = current_status, 
                        #         current_status_display = current_status_display,
                        #         previous_status = previous_status,
                        #         previous_status_display = previous_status_display,
                        #         stock_status_change_by_info = stock_status_change_by_info,
                        #         stock_in_date = approved_date,
                        #         created_by = created_by_qs
                        # )
                    
                    product_stock_list.append(product_stock_qs)
                    
                    # print(f"Barcode = {barcode}, Product Stock = {product_stock_qs}, {product_stock_list}")
                    
                    # print(f"......###...... This {barcode} All Barcode History Log Done Now Add In Stock Transfer ......###...... ")
                    
                    requisition_no = f"REG000{id}"
                    
                    product_stock_transfer_qs = ProductStockTransfer.objects.filter(requisition_no=requisition_no).last()
                    
                    stock_transfer_type = 'TRANSFER'
                        
                    if approval_status == 'True':
                        status = 'APPROVED'
                        
                    elif approval_status == 'False':
                        status = 'IN_TRANSIT'
                        
                    if not product_stock_transfer_qs:
                        
                        product_stock_transfer_qs = ProductStockTransfer.objects.create(
                            requisition_no=requisition_no, 
                            status = status, 
                            stock_transfer_type = stock_transfer_type, 
                            from_shop = stock_location_qs, 
                            to_shop=to_stock_location_qs, 
                            mismatch_barcode_list = None, 
                            not_received_barcode_list = None, 
                            received_barcode_list = received_sku_list, 
                            approved_by = employee_qs, 
                            remarks = f"Previous Stock Transfer is is {id}", 
                            created_by = created_by_qs, 
                            created_at = created_date
                            )
                    
                    elif product_stock_transfer_qs:
                        product_stock_transfer_qs = ProductStockTransfer.objects.filter(requisition_no=requisition_no)
                        
                        product_stock_transfer_qs = product_stock_transfer_qs.update(
                            requisition_no=requisition_no, 
                            status = status, 
                            stock_transfer_type = stock_transfer_type, 
                            from_shop = stock_location_qs, 
                            to_shop=to_stock_location_qs, 
                            mismatch_barcode_list = '', 
                            not_received_barcode_list = '', 
                            received_barcode_list = received_sku_list, 
                            approved_by = employee_qs, 
                            remarks = f"Previous Stock Transfer is is {id}", 
                            created_by = request.user, 
                            created_at = created_date
                            )
                        
                    print(f"......***...... This {requisition_no} Number is Generate ......***...... ")
                    
                    product_stock_transfer_qs = ProductStockTransfer.objects.filter(requisition_no=requisition_no).last()
                    
                    if product_stock_transfer_qs:
                        if product_stock_transfer_qs:
                            for i in product_stock_list:
                            
                                product_stock_transfer_qs.product_stock.add(i)
                        
                                # qs = ProductStock.objects.filter(barcode=i).last()
                                # print('ggggggggggggggg', product_stock_transfer_qs, requisition_no, i, qs)
                                # if qs:
                                #     # Assuming you need to add this product to a requisition related to the transfer
                                #     product_stock_transfer_qs.product_requisition.product_stock_set.add(qs)  # or the appropriate field/method
                                # else:
                                #     print(f'No ProductStock found for barcode {i}')
                        else:
                            print('product_stock_transfer_qs.product_requisition is None, cannot add qs')
                    else:
                        print('product_stock_transfer_qs is None, cannot proceed')
                        
                        
                    
                    product_stock_transfer_qs = ProductStockTransfer.objects.filter(requisition_no=requisition_no).last()
                    
                    if product_stock_transfer_qs.status == 'True':
                        # print('fffffffff', product_stock_transfer_qs.product_stock.values_list('barcode', flat = True))
                        new_sku_list__str = new_sku_list
                        product_stock_transfer_qs.received_barcode_list = str(new_sku_list__str)
                        
                        product_stock_transfer_qs.save()
                        
                    
                    product_stock_transfer_log_qs = ProductStockTransferLog.objects.filter(
                        product_stock = product_stock_transfer_qs, status = product_stock_transfer_qs.status
                        ).last()
                    
                    from_shop_info = OfficeLocationListSerializer(product_stock_transfer_qs.from_shop).data
                    to_shop_info = OfficeLocationListSerializer(product_stock_transfer_qs.to_shop).data
                    stock_status_change_by_info = BaseSerializer(product_stock_transfer_qs.created_by).data
                    
                    status = product_stock_transfer_qs.status
                    status_display = product_stock_transfer_qs.get_status_display()
                    
                    stock_transfer_type = product_stock_transfer_qs.stock_transfer_type
                    stock_transfer_type_display = product_stock_transfer_qs.get_stock_transfer_type_display()
                    
                    if not product_stock_transfer_log_qs:
                        product_stock_transfer_log_qs =  ProductStockTransferLog.objects.create(product_stock = product_stock_transfer_qs, from_shop_info = from_shop_info, to_shop_info = to_shop_info, status = status, status_display = status_display, stock_transfer_type = stock_transfer_type, stock_transfer_type_display = stock_transfer_type_display, stock_status_change_by_info = stock_status_change_by_info, status_changed_date = created_date, created_by = created_by_qs)
                        
                    elif product_stock_transfer_log_qs:
                        product_stock_transfer_log_qs = ProductStockTransferLog.objects.filter(product_stock = product_stock_transfer_qs)
                        
                        product_stock_transfer_log_qs =  product_stock_transfer_log_qs.update(product_stock = product_stock_transfer_qs, from_shop_info = from_shop_info, to_shop_info = to_shop_info, status = status, status_display = status_display, stock_transfer_type = stock_transfer_type, stock_transfer_type_display = stock_transfer_type_display, stock_status_change_by_info = stock_status_change_by_info, status_changed_date = created_date, created_by = created_by_qs)
                except:
                    pass
                
        qs = ProductStockTransfer.objects.all()
        
        return ResponseWrapper(msg=f"Total Stock Transfer Created = {qs.count()}",status=200)
            
            
    def bulk_discount_create(self, request,  *args, **kwargs):
        deal_of_the_week_list = DEALS_OF_THE_WEEK_LIST
        campaign_list = CAMPAIGN_LIST
        
        company_qs = Company.objects.filter().last()
        
        for item in deal_of_the_week_list:
            created_by = item.get('created_by').get('username')
            product_variant_name = item.get('product_variant').get('name')
            created_at = item.get('created_at')
            updated_at = item.get('updated_at')
            offer_price = item.get('offer_price')
            discount_value = item.get('discount_value')
            discount_amount = item.get('discount_amount')
            discount_type = item.get('discount_type')
            
            discount_name = '-'
            amount_type = '-'
            
            image = settings.NOT_FOUND_IMAGE
            
            if discount_type=='flat':
                amount_type = 'FLAT'
                discount_name = f"{discount_amount} TK. Off"
            else:
                amount_type = 'PERCENTAGE'
                discount_name = f"{discount_amount}% Off"
                
            # created_by_qs = UserAccount.objects.filter(email = created_by).last()
            # discount_qs = Discount.objects.filter(name = discount_name).last()
                
            created_by_qs = UserAccount.objects.filter().last()
            discount_qs = Discount.objects.filter(name = discount_name).last()
            
            product_price_info_qs = ProductPriceInfo.objects.filter(product__name = product_variant_name, product_price_type = 'ECOMMERCE').last()
            
            if not product_price_info_qs:
                print(f"{product_variant_name} is Not Found")
                
            if discount_qs:
                discount_qs = Discount.objects.filter(name = discount_name)
                
                qs = discount_qs.update(
                    name = discount_name,
                    image= image, 
                    is_for_lifetime = True,
                    discount_amount = discount_amount,
                    discount_status = 'DEAL_OF_WEEK',
                    company = company_qs,
                    created_at = created_at,
                    updated_at = updated_at, 
                    created_by = created_by_qs,
                    amount_type = amount_type,
                )
                
            else:
                slug = unique_slug_generator(name = discount_name)
                
                qs = Discount.objects.create(
                    name = discount_name,
                    slug= slug, 
                    image= image, 
                    is_for_lifetime = True,
                    discount_amount = discount_amount,
                    discount_status = 'DEAL_OF_WEEK',
                    company = company_qs,
                    created_at = created_at,
                    updated_at = updated_at, 
                    created_by = created_by_qs,
                    amount_type = amount_type,
                )
                
            discount_qs = Discount.objects.filter(name = discount_name).last()
            
            if product_price_info_qs:
                product_price_info_qs.discount = discount_qs
                product_price_info_qs.save()
                
            print(f"Deals Of The Weak, Name = {discount_name}")
        
        for item in campaign_list:
            created_by = item.get('created_by').get('username')
            # product_variant_name = item.get('product_variant').get('name')
            created_at = item.get('created_at')
            updated_at = item.get('updated_at')
            # offer_price = item.get('offer_price')
            # discount_value = item.get('discount_value')
            discount_amount = item.get('max_discount')
            discount_type = item.get('discount_type')
            name = item.get('name')
            slug = item.get('slug')
            image = item.get('image')
            description = item.get('description')
            terms_and_conditions = item.get('terms_and_conditions')
            
            start_date = item.get('start_date')
            end_date = item.get('end_date')
            
            discount_name = name
            slug = name
            
            amount_type = '-'
            
            # image = settings.NOT_FOUND_IMAGE
            
            if discount_type=='flat':
                amount_type = 'FLAT'
                # discount_name = f"{discount_amount} TK. Off"
            else:
                amount_type = 'PERCENTAGE'
                # discount_name = f"{discount_amount}% Off"
                
            created_by_qs = UserAccount.objects.filter().last()
            discount_qs = Discount.objects.filter(name = discount_name).last()
            
            # product_price_info_qs = ProductPriceInfo.objects.filter(product__name = product_variant_name, product_price_type = 'ECOMMERCE').last()
            
            # if not product_price_info_qs:
            #     print(f"{product_variant_name} is Not Found")
                
            if discount_qs:
                discount_qs = Discount.objects.filter(name = discount_name)
                
                qs = discount_qs.update(
                    name = discount_name,
                    slug= slug, 
                    image= image, 
                    description =description,
                    discount_amount = discount_amount,
                    discount_status = 'CAMPAIGN',
                    company = company_qs,
                    created_at = created_at,
                    updated_at = updated_at, 
                    created_by = created_by_qs,
                    amount_type = amount_type,
                    terms_and_conditions=terms_and_conditions, 
                    start_date = start_date,
                    end_date = end_date,
                )
                
            else:
                # slug = unique_slug_generator(name = discount_name)
                
                qs = Discount.objects.create(
                    name = discount_name,
                    slug= slug, 
                    image= image, 
                    description =description,
                    discount_amount = discount_amount,
                    discount_status = 'CAMPAIGN',
                    company = company_qs,
                    created_at = created_at,
                    updated_at = updated_at, 
                    created_by = created_by_qs,
                    amount_type = amount_type,
                    terms_and_conditions=terms_and_conditions, 
                    start_date = start_date,
                    end_date = end_date,
                )
                
            discount_qs = Discount.objects.filter(name = discount_name).last()
            
            # if product_price_info_qs:
            #     product_price_info_qs.discount = discount_qs
            #     product_price_info_qs.save()
                
            print(f"Campaigns , Name = {discount_name}")
                
        
        return ResponseWrapper(msg='Success', status=200)
    
    def bulk_shop_day_end_create(self, request,  *args, **kwargs):
        day_end_list = DAY_END_LIST 
        
        company_qs = Company.objects.filter().last()
        
        for item in day_end_list:
            store_name = item.get('store_name')
            day_end_time = item.get('day_end_time')
            total_shop_sale = item.get('total_shop_sale')
            total_ecommerce_collection = item.get('total_ecommerce_collection')
            total_nazat_collection = item.get('total_prothoma_collection')
            petty_cash = item.get('petty_cash')
            opening_amount = item.get('opening_amount')
            total_bank_deposit = item.get('total_bank_deposit')
            total_expenses = item.get('total_expenses')
            balance = round(item.get('balance'), 2)
            note = item.get('note')
            mfs = item.get('mfs')
            b2b = item.get('b2b')
            warranty_claim_qty = item.get('warranty_claim_qty')
            gsheba_claim_qty = item.get('gsheba_claim_qty')
            refund_amount = item.get('refund_amount')
            
            store_qs = OfficeLocation.objects.filter(name__icontains = store_name).last()
            
            # created_at_qs = UserAccount.objects.filter(name__icontains = store_name).last()
            
            
                    
            day_end_time_str = day_end_time
            
            day_end_time = datetime.fromisoformat(day_end_time_str)
            
            day_end_qs = ShopDayEnd.objects.filter(
                shop = store_qs, day_end_date__date = day_end_time.date()
            ).last()
            
            panel_partnership = [
                {
                    "name": "Nazat", 
                    "total_sell_amount": total_nazat_collection,
                    "total_gsheba_amount": 0
                }
            ]
            print(f'day_end_qs = {day_end_qs}, day_end_time.date()= {day_end_time.date()}')
            
            currency_collection = [{"currency": "1000", "quantity": 0}, {"currency": "500", "quantity": 0}, {"currency": "200", "quantity": 0}, {"currency": "100", "quantity": 0}, {"currency": "50", "quantity": 0}, {"currency": "20", "quantity": 0}, {"currency": "10", "quantity": 0}, {"currency": "5", "quantity": 0}, {"currency": "2", "quantity": 0}, {"currency": "1", "quantity": 0}]
            
            mfs_collection = MFS_COLLECTION
            
            qs = day_end_qs
            
            if day_end_qs:
                
                if not day_end_qs.slug:
                    slug = unique_slug_generator(name = store_name)
                elif day_end_qs.slug:
                    slug = day_end_qs.slug

                # Filter ShopDayEnd objects
                day_end_qs = ShopDayEnd.objects.filter(
                    shop=store_qs, day_end_date__date=day_end_time.date()
                )
                
                day_end_qs = day_end_qs.update(
                    shop = store_qs,
                    slug = slug,
                    retail_sell_amount = total_shop_sale,
                    panel_partnership = panel_partnership,
                    e_retail_sell_amount = 0.0,
                    ecommerce_collection_amount = total_ecommerce_collection, 
                    refund_amount = refund_amount,
                    warranty_claim_quantity = warranty_claim_qty,
                    gsheba_claim_quantity = gsheba_claim_qty,
                    total_b2b_sell_amount = b2b, 
                    currency_collection = currency_collection,
                    total_bank_deposit_amount = total_bank_deposit,
                    total_expense_amount = total_expenses,
                    petty_cash_amount = petty_cash,
                    opening_balance_amount = opening_amount,
                    current_balance_amount = balance,
                    remarks = note,
                    day_end_date = day_end_time.date(),
                    created_at = day_end_time.date(),
                    created_by = request.user,
                    mfs_collection = mfs_collection
                )
                
            if not day_end_qs:
                slug = unique_slug_generator(name = store_name)
                
                day_end_qs = ShopDayEnd.objects.create(
                    shop = store_qs,
                    slug = slug,
                    retail_sell_amount = total_shop_sale,
                    panel_partnership = panel_partnership,
                    e_retail_sell_amount = 0.0,
                    ecommerce_collection_amount = total_ecommerce_collection, 
                    refund_amount = refund_amount,
                    warranty_claim_quantity = warranty_claim_qty,
                    gsheba_claim_quantity = gsheba_claim_qty,
                    total_b2b_sell_amount = b2b, 
                    currency_collection = currency_collection,
                    total_bank_deposit_amount = total_bank_deposit,
                    total_expense_amount = total_expenses,
                    petty_cash_amount = petty_cash,
                    opening_balance_amount = opening_amount,
                    current_balance_amount = balance,
                    remarks = note,
                    day_end_date = day_end_time.date(),
                    created_at = day_end_time.date(),
                    created_by = request.user,
                    mfs_collection = mfs_collection
                )
                
            day_end_qs = ShopDayEnd.objects.filter(
                shop = store_qs, day_end_date__date = day_end_time.date()
            ).last()
            
            day_end_qs.day_end_date = day_end_time.date()
            day_end_qs.created_at = day_end_time.date()
            day_end_qs.save()
                
            print(f"ID = {day_end_qs.id}, Name = {store_name}, Day End = {day_end_time.date()}, Save Data is = {day_end_qs.day_end_date}")
        
        return ResponseWrapper(msg='Success', status=200)
    
    
    @log_activity
    def bulk_employee_information_create(self, request,  *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        count = 1
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            si = row[0]
            name = row[1]
            designation = row[2]
            ranking = row[3]
            department = row[4]
            employee_type = row[5]
            work_station = row[6]
            joining_date = row[9]
            employee_id = row[10]
            status = row[11]
            last_working_date = row[14]
            birth_date = row[15]
            email = row[17]
            phone = row[18]
            
            if not si:
                break
            
            print(f"CCCCCCCCC= {row_number}, name = {name}, Employee id  ={employee_id}")
            
            employee_id = employee_id
            email = employee_id
            phone = employee_id
            
            if not name:
                print('gggggggggggg')
            
            else:
                employee_qs = EmployeeInformation.objects.filter(
                    Q(employee_id = employee_id)
                    | Q(user__email = employee_id)
                    | Q(user__phone = employee_id)
                    ).last()
                
                if not employee_qs:
                    
                    print(f"Row Number = {row_number}, name = {name}")
                    
                    user_qs= UserAccount.objects.filter(
                        phone = phone
                    ).last()
                    
                    employee_qs = EmployeeInformation.objects.filter(user = user_qs, employee_id = employee_id).last()
                    
                    count +=1
                    
                    today = settings.TODAY
                    
                    slug = unique_slug_generator(name = f"{name}-{employee_id}")
                    
                    if employee_qs:
                        employee_qs.name = name
                        employee_qs.save()
                    
                    
                    country_qs = Country.objects.filter(name = 'Bangladesh').last()
                    
                    company_qs = Company.objects.filter().last()
                    
                    department_qs = Department.objects.filter(name = department).last()
                    
                    if department_qs:
                        department_qs.name = department
                        department_qs.save()
                    
                    elif not department_qs:
                        slug = unique_slug_generator(name = department)
                        
                        department_qs = Department.objects.create(
                            name = department, slug = slug, created_by = request.user
                        ) 
                    
                    designation_qs = Designation.objects.filter(name = designation).last()
                    
                    
                    if designation_qs:
                        designation_qs.name = designation
                        designation_qs.departments = department_qs
                        designation_qs.save()
                    
                    elif not designation_qs:
                        slug = unique_slug_generator(name = designation)
                        
                        designation_qs = Designation.objects.create(
                            name = designation, slug = slug, created_by = request.user,
                            departments = department_qs
                        )
                    
                    ranking_qs = None
                    
                    try:
                        if ranking:   
                            ranking_qs = Ranking.objects.filter(name = ranking).last()
                            
                        else:
                            ranking_qs = Ranking.objects.all().last()
                        
                        
                        if ranking_qs:
                            ranking_qs.name = ranking
                            ranking_qs.save()
                        
                        elif not ranking_qs:
                            slug = unique_slug_generator(name = ranking)
                            
                            ranking_qs = Ranking.objects.create(
                                name = ranking, slug = slug, created_by = request.user 
                            )
                            
                    except:
                        pass
                    
                    print(f"Employee Type = {employee_type}")
                    
                    employee_type_qs = EmployeeType.objects.filter(name = employee_type).last()
                    
                    
                    if employee_type_qs:
                        employee_type_qs.name = employee_type
                        employee_type_qs.save()
                    
                    elif not employee_type_qs:
                        slug = unique_slug_generator(name = employee_type)
                        
                        employee_type_qs = EmployeeType.objects.create(
                            name = employee_type, slug = slug, created_by = request.user
                        ) 
                        
                    work_station_qs = None
                    
                    try:
                        work_station_qs = OfficeLocation.objects.filter(name__icontains = work_station).last()
                    
                        if not work_station_qs:
                            work_station_qs = OfficeLocation.objects.filter(name__icontains = "Head Office").last()
                        
                        print("Office Location = {work_station_qs.name}")
                        
                    except:
                        pass
                    
                    image_url = settings.NOT_FOUND_IMAGE
                    
                    if not phone:
                        phone = employee_id
                        
                    if not email:
                        email = employee_id
                    
                    user_qs = UserAccount.objects.filter(
                        Q(email = email)
                        | Q(phone = phone)
                        
                    )
                    
                    if user_qs:
                        email = employee_id
                        
                    is_active = False
                    if status == "Active":
                        is_active = True
                        
                    print(f"Row Number = {row_number} and Email = {email} Phone = {phone}")
                    
                    password ="GPRO@1234"
                    
                    password = make_password(password=password)
                        
                        
                    parts = name.split()
                    
                    first_name = parts[0]
                        
                    last_name = ' '.join(parts[1:])
                        
                    print(f"First Name = {first_name}, and Last Name = {last_name} and Employee Id = {employee_id}")
                    
                    if not joining_date:
                        joining_date = "2018-01-01"
                        
                    if not user_qs:
                        qs = UserAccount.objects.create(
                            password=password,
                            first_name=first_name,
                            last_name=last_name,
                            email=email,
                            phone=phone,
                            is_active=is_active,
                            date_joined=joining_date,
                            is_staff=True,
                            is_superuser=False,
                        )
                        
                    else:
                        qs = user_qs.update(
                            # password=password,
                            first_name=first_name,
                            last_name=last_name,
                            email=email,
                            phone=phone,
                            is_active=is_active,
                            date_joined=joining_date,
                            is_staff=True,
                            is_superuser=False,
                        )
                        
                    user_qs = User.objects.filter(email=email,
                            phone=phone).last()
                    
                    
                    user_type = 'Employee'
                    
                    user_type_qs = UserType.objects.filter(name = "Employee").last()
                    
                    user_group_qs = UserGroup.objects.filter(name = user_type).last()
                    
                    if not user_group_qs:
                        user_group_qs = UserGroup.objects.create(name__icontains = "Employee")
                    
                    if user_group_qs:
                        user_qs.groups.add(user_group_qs)
                    
                    if not user_type_qs:
                        slug = unique_slug_generator(name = user_type)
                        
                        user_type_qs = UserType.objects.create(name = user_type, slug = slug, created_by = request.user)
                        
                    name = f"{user_qs.first_name} {user_qs.last_name}"
                    
                    user_info_qs = UserInformation.objects.filter(
                        user_id=user_qs.id
                    )
                    if not user_info_qs:
                        slug = unique_slug_generator(name = name)
                        
                        user_info_qs = UserInformation.objects.create(
                            name=name, user_id=user_qs.id, image = image_url, created_by = request.user
                        )
                    else:
                        user_info_qs = user_info_qs.update(
                            name=name, image = image_url
                        )
                        
                    user_information_qs = UserInformation.objects.filter(
                        user_id=user_qs.id
                    ).last()
                        
                    if user_type_qs:
                        user_information_qs.user_type = user_type_qs
                        user_information_qs.save()
                        
                    employee_qs = EmployeeInformation.objects.filter(user = user_qs, employee_id = employee_id).last()
                    
                    count +=1
                    
                    today = settings.TODAY
                    
                    slug = unique_slug_generator(name = f"{name}-{employee_id}")
                    
                    if employee_qs:
                        employee_qs.name = name
                        employee_qs.save()
                    
                    try:
                        if not employee_qs:
                        
                            employee_qs = EmployeeInformation.objects.create(user = user_qs, phone_number=phone, employee_id = employee_id, image = image_url, name = name, slug = slug, employee_company = company_qs, joining_date = joining_date, date_of_birth = birth_date,employee_type = employee_type_qs, designations = designation_qs, work_station = work_station_qs, rank = ranking_qs, created_by = request.user) 
                        
                            if employee_qs:
                                
                                employee_qs = EmployeeInformation.objects.filter(user = user_qs)
                                
                                employee_qs = employee_qs.update(user = user_qs, phone_number=phone, employee_id = employee_id, image = image_url, name = name, slug = slug, employee_company = company_qs, joining_date = joining_date, date_of_birth = birth_date,employee_type = employee_type_qs, designations = designation_qs, work_station = work_station_qs, rank = ranking_qs, created_by = request.user)
                    except:
                        pass
                    
                print(f"After = {row_number}, name = {name}, Employee id  ={employee_id}, Employee User Id = {employee_qs}")
                
        print('Employee Create Done')
                
        return ResponseWrapper(msg='Success', status=200)
    
    
    
    
    def bulk_order_create(self, request, *args, **kwargs):
        company_qs = Company.objects.filter().last()
        
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        count = 19990
        total_rows = 23910  # Assuming the total number of rows is 123, adjust if needed
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=19991, values_only=True), start=1):
            rest_of = total_rows - count
            
            id = row[0]
            order_date = row[1]
            invoice_no = row[2]
            customer_name = row[3]
            mobile_number = row[4]
            email = row[5]
            address = row[6]
            courier_name = row[7]
            store_name = row[8]
            order_status = row[9]
            order_type = row[10]
            payment_type = row[11]
            note = row[12]
            payment_status = row[13]
            delivery_type = row[14]
            delivery_method = row[15]
            product_name = row[16]
            brand = row[17]
            seller = row[18]
            product_code = row[19]
            sku_list = row[20]
            quantity = row[21]
            msp = row[22]
            mrp = row[23]
            selling_price = row[24]
            promo_discount_amount = row[25]
            after_promo_selling_price = row[26]
            promo_code = row[27]
            commission_amount = row[28]
            total_product_price = row[29]
            total_discount_amount = row[30]
            total_net_payable_amount = row[31]
            total_gsheba_amount = row[32]
            total_delivery_charge = row[33]
            total_tax_amount = row[34]
            total_promo_discount = row[35]
            total_advance_amount = row[36]
            total_due_amount = row[37]
            total_payable_amount = row[38]
            
            if not id:
                break
            
            print(f"SI = {count}, Total = {total_rows}, Rest Of = {rest_of}, Invoice No  = {invoice_no}")
            count += 1
            
            order_qs = Order.objects.filter(invoice_no = invoice_no).last()
            
            store_name_qs = None
            area_qs = None
            pickup_shop_qs = None
            user_qs = None
            delivery_method_qs = None
            customer_qs = None
            
            area_name = "-"
            district_name = "-"
            division_name = "-"
            country_name = "Bangladesh"
            
            
            # Order Type Start
            
            if order_type == "Point of Sell":
                order_type = "POINT_OF_SELL"
                
            elif order_type == "E-Commerce Sell":
                order_type = "ECOMMERCE_SELL"
                
            elif order_type == "E-Commerce Sell":
                order_type = "ECOMMERCE_SELL"
                
            else:
                order_type = "RETAIL_ECOMMERCE_SELL"
            
            # Order Type End
            
            # Order Status Start
            
            if order_status == "Delivered":
                order_status = "DELIVERED"
            elif order_status == "Cancelled":
                order_status = "CANCELLED"
            else:
                try:
                    order_status = order_status.replace(" ", "_").upper()
                except:
                    order_status = "DELIVERED"
                    
                # order_status = "ORDER_CONFIRMED"
            
            # Order Status End
            
            # Order Delivery Method Start
            
            if delivery_method == "store_sale":
                delivery_method = "STORE_SELL"
                
            elif delivery_method == "store_pickup":
                delivery_method = "SHOP_PICKUP"
                
            delivery_method_qs = DeliveryMethod.objects.filter(delivery_type = delivery_method).last()
                
            # else:
            #     delivery_method = "ORDER_CONFIRMED"
            
            # Order Delivery Method End
            
            approved_status = "APPROVED"
            
        
            if payment_status == "PAID":
                payment_status = "PAID"
                
            elif payment_status == "UNPAID":
                payment_status = "UNPAID"
            
            if payment_type == "Cash on Delivery":
                payment_type = "CASH_ON_DELIVERY"
                
            elif payment_type == "Online Payment":
                payment_type = "ONLINE_PAYMENT"
            
                
            store_name_qs = OfficeLocation.objects.filter(name = store_name).last()
            
            if store_name_qs:
                try:
                    area_qs = store_name_qs.area
                except:
                    area_qs = Area.objects.filter(name__icontains = "Dhaka").last()
                
            if order_type == "POINT_OF_SELL":
                pickup_shop_qs = store_name_qs
                
            if area_qs:
                area_name = area_qs.name
                district_name = area_qs.district.name
                division_name = area_qs.district.division.name
                country_name = "Bangladesh"
                
            address = address
            
            customer_qs = UserInformation.objects.filter(
                Q(user__phone = mobile_number)
                |Q(user__email = email)
            ).last()
            
            # print('ggggggggggg', customer_qs)
            
            if not customer_name:
                customer_name = "Walk in Way Customer"
            
            if not customer_qs:
                # print(f"Line 3996")
                
                user_qs = UserAccount.objects.filter(
                    Q(phone = mobile_number)
                    |Q(email = email)
                ).last()
                
                if not user_qs:  
                    user_qs = UserAccount.objects.create(
                        first_name = customer_name,
                        last_name = customer_name,
                        email = email,
                        phone = mobile_number
                    )
                    
                if user_qs:
                    user_qs.first_name = customer_name
                    user_qs.last_name = customer_name
                    user_qs.email = email
                    user_qs.phone = mobile_number
                    user_qs.save()
                    
                user_type_qs = UserType.objects.filter(
                    name__icontains = "Customer"
                ).last()
                
                customer_qs = UserInformation.objects.filter(user = user_qs).last()
                
                if not customer_qs:
                    
                    slug = unique_slug_generator(name= f"walk-in-way-customer-{count}")
                    
                    customer_qs = UserInformation.objects.create(
                        user = user_qs,
                        name = customer_name,
                        slug = slug,
                        address = address,
                        image = settings.NOT_FOUND_IMAGE,
                        user_type = user_type_qs,
                        created_by = request.user
                    )
                # print(f"Line 4037")
            
            elif customer_qs:
                # customer_qs = UserInformation.objects.filter(user = user_qs).last()
                # print(f"Line 4041")
                
                try:
                    if not customer_name:
                        # print(f"Line 4045")
                        customer_name = "Walk in Way Customer"
                    
                    if customer_qs:
                        # print(f"Line 4049")
                        customer_qs.name = customer_name
                        customer_qs.save()
                except:
                    pass
                
            customer_qs = UserInformation.objects.filter(name = customer_name).last()
            
            if not customer_qs:
                customer_qs = UserInformation.objects.filter(name = "Walk in Way Customer").last()
                
            # customer_qs = UserInformation.objects.filter(user = user_qs).last()
            
            # print('Customer Name =', customer_qs.name)
            
            if order_qs:
                # print(f"{invoice_no} Order Update")
                qs = Order.objects.filter(invoice_no = invoice_no)
                
                qs = qs.update(
                    invoice_no = invoice_no,
                    order_type = order_type,
                    status = order_status,
                    approved_status = approved_status,
                    delivery_method = delivery_method_qs,
                    payment_status = payment_status,
                    payment_type = payment_type,
                    area = area_qs,
                    shop = store_name_qs,
                    pickup_shop = pickup_shop_qs,
                    customer = customer_qs,
                    district_name = district_name,
                    division_name = division_name,
                    country_name = country_name,
                    address = address,
                    remarks = note,
                    promo_code = promo_code,
                    total_product_price = total_product_price,
                    total_discount_amount = total_discount_amount,
                    total_net_payable_amount = total_net_payable_amount,
                    total_gsheba_amount = total_gsheba_amount,
                    total_tax_amount = total_tax_amount,
                    total_promo_discount = total_promo_discount,
                    total_delivery_charge = total_delivery_charge,
                    total_payable_amount = total_payable_amount,
                    total_advance_amount = total_advance_amount,
                    total_paid_amount = total_advance_amount,
                    total_due_amount = total_due_amount,
                    order_date = order_date,
                    created_at = order_date,
                    created_by = request.user
                )
                
            
            if not order_qs:
                # print(f"{invoice_no} Order Create")
                
                order_qs = Order.objects.create(
                    invoice_no = invoice_no,
                    order_type = order_type,
                    status = order_status,
                    approved_status = approved_status,
                    delivery_method = delivery_method_qs,
                    payment_status = payment_status,
                    payment_type = payment_type,
                    area = area_qs,
                    shop = store_name_qs,
                    customer = customer_qs,
                    pickup_shop = pickup_shop_qs,
                    district_name = district_name,
                    division_name = division_name,
                    country_name = country_name,
                    address = address,
                    remarks = note,
                    promo_code = promo_code,
                    total_product_price = total_product_price,
                    total_discount_amount = total_discount_amount,
                    total_net_payable_amount = total_net_payable_amount,
                    total_gsheba_amount = total_gsheba_amount,
                    total_tax_amount = total_tax_amount,
                    total_promo_discount = total_promo_discount,
                    total_delivery_charge = total_delivery_charge,
                    total_payable_amount = total_payable_amount,
                    total_advance_amount = total_advance_amount,
                    total_paid_amount = total_advance_amount,
                    total_due_amount = total_due_amount,
                    # total_promo_discount = promo_discount_amount,
                    order_date = order_date,
                    created_at = order_date,
                    created_by = request.user,
                )
                
            order_qs = Order.objects.filter(invoice_no = invoice_no).last()
            
            # print(f'Customer Name = {order_qs.customer},{customer_qs}')
            
            customer_address_log_qs = CustomerAddressInfoLog.objects.filter(
                order = order_qs
            ).last()
            
            address_type = "HOME"
            
            if customer_address_log_qs:
                customer_address_log_qs.order = order_qs
                customer_address_log_qs.address_type = address_type
                customer_address_log_qs.name = customer_qs.name
                customer_address_log_qs.phone = customer_qs.user.phone
                customer_address_log_qs.email = customer_qs.user.email
                customer_address_log_qs.address = address
                customer_address_log_qs.area_name = area_name
                customer_address_log_qs.district_name = district_name
                customer_address_log_qs.division_name = division_name
                customer_address_log_qs.country_name = country_name
                
                customer_address_log_qs.save()
                
            if not customer_address_log_qs:
                slug = unique_slug_generator(name = order_qs.customer.name)
                
                customer_address_log_qs = CustomerAddressInfoLog.objects.create(
                    order = order_qs,
                    slug = slug,
                    address_type = address_type,
                    name = customer_qs.name,
                    phone = customer_qs.user.phone,
                    email = customer_qs.user.email,
                    address = address,
                    area_name = area_name,
                    district_name = district_name,
                    division_name = division_name,
                    country_name = country_name,
                    created_by = request.user,
                )
            
            
            # Order Payment Log
            
            customer_payment_log_qs = OrderPaymentLog.objects.filter(
                order = order_qs
            ).last()
            
            payment_status = "RECEIVED"
            
            order_payment_qs = PaymentType.objects.filter(
                name__icontains ="Cash On Delivery"
            ).last()
            
            if customer_payment_log_qs:
                customer_payment_log_qs.order = order_qs
                customer_payment_log_qs.order_payment = order_payment_qs
                customer_payment_log_qs.received_amount = total_payable_amount
                customer_payment_log_qs.created_at = order_qs.order_date
                customer_address_log_qs.save()
                
            if not customer_payment_log_qs:
                slug = unique_slug_generator(name = order_qs.customer.name)
                
                customer_payment_log_qs = OrderPaymentLog.objects.create(
                    order = order_qs,
                    slug = slug,
                    order_payment = order_payment_qs,
                    received_amount = total_payable_amount,
                    created_at = order_qs.order_date,
                    created_by = request.user,
                )
                
            order_item_qs = OrderItem.objects.filter(
                order = order_qs
            ).last()
            
            product_qs = Product.objects.filter(
                name = product_name
            ).last()
            
            # sku_list = list(set(sku_list))  # Ensure there are no duplicates
            # formatted_sku_list = []

            # for sku in sku_list:
            #     for char in sku:
            #         formatted_sku_list.append(f'[[\\"{char}\\"]]')

            # barcode = ', '.join(formatted_sku_list)
            
            # barcode = next((barcode for barcode in sku_list if barcode.startswith(product_code)))
            
            barcode_list = ast.literal_eval(sku_list)
            
            barcode = next((barcode for barcode in barcode_list if barcode.startswith(product_code)), "-")
            
            # print(f"Barcode = {barcode}, product_code={product_code} sku  = {sku_list}")
            
            try:
                product_stock_qs = ProductStock.objects.filter(barcode = barcode).last()
                
                stock_location_qs = order_qs.shop
                # print(f"SHop = {order_qs.shop}")
                
                date = order_qs.order_date.date() + timedelta(days=1)
                
                remarks = f"This Barcode History is Generated By MIS Department at {date} and Order Invoice No is {invoice_no}, Also Sold From {stock_location_qs.name}"
                
                if product_stock_qs:
                    # print(f"{barcode} is Update ....")
                    
                    product_stock_qs.status = "SOLD"
                    product_stock_qs.stock_location = stock_location_qs
                    product_stock_qs.created_at = order_qs.order_date
                    product_stock_qs.created_by = order_qs.created_by
                    product_stock_qs.remarks = remarks
                    product_stock_qs.save()
                    
                    product_stock_log_qs = ProductStockLog.objects.filter(
                    product_stock = product_stock_qs,
                    current_status = "SOLD"
                    ).last()
                    
                    # print(f"ffff= {product_stock_log_qs}")
                    
                    if not product_stock_log_qs:
                        barcode_status_log(product_stock_qs = product_stock_qs, previous_status = 'ACTIVE', previous_status_display ='Active'  , current_status = "SOLD", remarks = product_stock_qs.remarks, is_active = True, request_user = order_qs.created_by,stock_in_date = order_qs.order_date)
                        
                    else:
                        product_stock_log_qs.remarks = remarks
                        product_stock_log_qs.created_at = order_qs.order_date
                        product_stock_log_qs.save()
                        
                        
                else:
                    # print(f"{barcode} is Creating ....")
                    product_price_info_qs = ProductPriceInfo.objects.filter(
                        product__slug = product_qs.slug,
                        product_price_type = "POINT_OF_SELL"
                    ).last()
                    
                    product_stock_qs = ProductStock.objects.create(
                        barcode = barcode,
                        product_price_info = product_price_info_qs,
                        status = "SOLD",
                        stock_location = stock_location_qs,
                        stock_in_date = order_qs.order_date,
                        created_at = order_qs.order_date,
                        created_by = order_qs.created_by,
                        remarks = remarks
                        )
                    
                    
                    product_stock_log_qs = ProductStockLog.objects.filter(
                    product_stock = product_stock_qs,
                    current_status = "SOLD"
                    ).last()
                    
                    # print(f"ffff= {product_stock_log_qs}")
                    
                    if not product_stock_log_qs:
                        barcode_status_log(product_stock_qs = product_stock_qs, previous_status = 'ACTIVE', previous_status_display ='Active'  , current_status = "SOLD", remarks = product_stock_qs.remarks, is_active = True, request_user = order_qs.created_by,stock_in_date = order_qs.order_date)
                        
                    else:
                        product_stock_log_qs.remarks = remarks
                        product_stock_log_qs.created_at = order_qs.order_date
                        product_stock_log_qs.save()
                
                
            except:
                pass
            

            try:
                if order_item_qs:
                    order_item_qs = OrderItem.objects.filter(
                        order = order_qs
                    )
                    qs = order_item_qs.update(
                        order = order_qs,
                        # status = order_qs.status,
                        quantity = quantity,
                        product = product_qs,
                        unit_msp_price = msp,
                        unit_mrp_price = mrp,
                        selling_price = selling_price,
                        total_product_price = (total_product_price*quantity),
                        total_tax_amount = total_tax_amount,
                        commission_amount = commission_amount,
                        barcode_number = barcode,
                        remarks = str(set(sku_list)),
                        promo_code = promo_code,
                        gsheba_amount = total_gsheba_amount,
                        
                        created_by = request.user,
                        created_at = order_date
                    )
                    
                if not order_item_qs:
                    
                    order_item_qs = OrderItem.objects.create(
                        order = order_qs,
                        status = order_qs.status,
                        quantity = quantity,
                        product = product_qs,
                        unit_msp_price = msp,
                        unit_mrp_price = mrp,
                        selling_price = selling_price,
                        total_product_price = (total_product_price*quantity),
                        total_tax_amount = total_tax_amount,
                        commission_amount = commission_amount,
                        barcode_number = barcode,
                        remarks = str(set(sku_list)),
                        promo_code = promo_code,
                        gsheba_amount = total_gsheba_amount,
                        
                        created_by = request.user,
                        created_at = order_date
                    )
                    
                if order_qs.status == "DELIVERED":
                    
                    order_item_qs = OrderItem.objects.filter(
                        order = order_qs
                    ).last()
                            
                    product_warranty_list = product_qs.product_warrantys.all().order_by('warranty_type')
                    
                    if product_warranty_list:
                        for product_warranty in product_warranty_list:
                            warranty_type = product_warranty.warranty_type
                            warranty_duration = product_warranty.warranty_duration
                            value = product_warranty.value
                            
                            start_date = order_qs.order_date
                            
                            if warranty_type == "1_GSHEBA_WARRANTY" and order_item_qs.gsheba_amount == 0.0:
                                pass
                            else:
                                order_item_warranty_qs = order_item_qs.order_item_warranty_logs.all().last()
                                
                                if order_item_warranty_qs:
                                    start_date = order_qs.order_date
                                    
                                if warranty_duration == 'DAY':
                                    end_date = start_date + timedelta(days=int(value))
                                elif warranty_duration == 'MONTH':
                                    end_date = start_date + timedelta(days=30 * int(value))  # Assuming 1 month = 30 days
                                elif warranty_duration == 'YEAR':
                                    end_date = start_date + timedelta(days=365 * int(value)) 
                                    
                                # print(f"Warranty Start Date = {start_date} & End Date = {end_date}")
                                
                                if order_item_warranty_qs:
                                    # print('AAAA')
                                    order_item_warranty_log_qs = order_item_qs.order_item_warranty_logs.filter(warranty_type = warranty_type)
                                    
                                    if not order_item_warranty_log_qs:
                                        # print('BBBB')
                                        order_item_warranty_qs = OrderItemWarrantyLog.objects.create(
                                            order_item = order_item_qs,
                                            warranty_type = warranty_type,
                                            warranty_duration = warranty_duration,
                                            value = value,
                                            start_date = start_date,
                                            end_date = end_date,
                                            created_by = order_qs.created_by
                                        )
                                    else:
                                        # print('DDDD')
                                        
                                        order_item_warranty_log_qs.update(
                                            order_item = order_item_qs,
                                            warranty_type = warranty_type,
                                            warranty_duration = warranty_duration,
                                            value = value,
                                            start_date = start_date,
                                            end_date = end_date,
                                            created_by = order_qs.created_by
                                        )
                                        
                                # elif not order_item_warranty_qs:
                                    # print('CCC')
                                    order_item_warranty_qs = OrderItemWarrantyLog.objects.create(
                                            order_item = order_item_qs,
                                            warranty_type = warranty_type,
                                            warranty_duration = warranty_duration,
                                            value = value,
                                            start_date = start_date,
                                            end_date = end_date,
                                            created_by = order_qs.created_by
                                        )
                                    
                                # print(f"After Warranty Start Date = {order_item_warranty_qs.start_date} & End Date = {order_item_warranty_qs.end_date}")
                        
            except:
                pass

        # print('Order Create Done')
                
        return ResponseWrapper(msg='Success', status=200) 
    
    
    @log_activity
    def bulk_shop_user_create(self, request,  *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        count = 1
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            si = row[0]
            name = row[1]
            email = row[2]
            password = row[3]
            
            phone = email
            
            if not email:
                break
            
            user_qs = UserAccount.objects.filter(
                email = email
            )
            
            first_name = "GProjukti"
            
            last_name = email.split('@')[0].split('.')[-1].capitalize()
            
            # store_name = email.split('_')[0].split('.')[-1].capitalize()
            store_name = "GProjukti.com - Test Shop"
            
            name = f"{first_name} {last_name}"
            
            image_url = settings.NOT_FOUND_IMAGE
            
            password = make_password(password)
                
            if not user_qs:
                qs = UserAccount.objects.create(
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    phone=phone,
                    is_active=True,
                    is_staff=False,
                    is_superuser=False
                )
                
            else:
                qs = user_qs.update(
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    phone=phone,
                    is_active=True,
                    is_staff=False,
                    is_superuser=False
                )
                
            user_qs = User.objects.filter(email=email,
                    phone=phone).last()
            
            user_type_qs = UserType.objects.filter(name__icontains = "Shop").last()
            
            user_group_qs = UserGroup.objects.filter(name__icontains = "Shop").last()
            
            if not user_group_qs:
                user_group_qs = UserGroup.objects.create(name__icontains = "Shop")
            
            if user_group_qs:
                user_qs.groups.add(user_group_qs)
            
            if not user_type_qs:
                slug = unique_slug_generator(name__icontains = "Shop")
                
                user_type_qs = UserType.objects.create(name__icontains = "Shop", slug = slug, created_by = request.user)
                
            
            user_info_qs = UserInformation.objects.filter(
                user_id=user_qs.id
            )
            if not user_info_qs:
                slug = unique_slug_generator(name = name)
                
                user_info_qs = UserInformation.objects.create(
                    name=name, user_id=user_qs.id, image = image_url, created_by = request.user
                )
            else:
                user_info_qs = user_info_qs.update(
                    name=name, image = image_url
                )
                
            user_information_qs = UserInformation.objects.filter(
                user_id=user_qs.id
            ).last()
                
            if user_type_qs:
                user_information_qs.user_type = user_type_qs
                user_information_qs.save()
            
            
            employee_qs = EmployeeInformation.objects.filter(user = user_qs).last()
            employee_id = f"GPRO-4000{count}"
            
            count +=1
            company_qs = Company.objects.all().last()
            
            today = settings.TODAY
            
            print(f"User Name = {name}")
            
            slug = unique_slug_generator(name = name)
            
            if not employee_qs:
                qs = EmployeeInformation.objects.create(user = user_qs, employee_id = employee_id, image = image_url, name = name, slug = slug, employee_company = company_qs, joining_date = today, date_of_birth = today, created_by = request.user) 
            
            if employee_qs:
                qs = EmployeeInformation.objects.filter(user = user_qs)
                
                employee_qs = qs.update(user = user_qs, employee_id = employee_id, image = image_url, name = name, slug = slug, employee_company = company_qs, joining_date = today, date_of_birth = today, created_by = request.user)
                
            work_station_qs = OfficeLocation.objects.filter(
                name__icontains = store_name
            ).last()
            
            employee_qs = EmployeeInformation.objects.filter(user = user_qs).last()
            
            if work_station_qs:
                # print(f"Work Station Name = {work_station_qs.name}")
                employee_qs.work_station = work_station_qs
                employee_qs.save()
            
            if not work_station_qs:
                print(f"Not Found Work Station Name = {store_name}")
            
                
        return ResponseWrapper(msg='Success', status=200) 
            

    def multiple_user_permission_create(self, request,  *args, **kwargs):
        user_group_list = USER_GROUP_PERMISSION
        
        for item in user_group_list:
            group_name = item.get('group_name')
            
            custom_permission_list = item.get('custom_permission_list')
            
            print(f'................{group_name}......................') 
            
            qs = UserGroup.objects.filter(name__icontains = group_name).last()
            
            if qs:
                print(f"Update = {group_name}")
                qs.name =group_name
                qs.save()
                
            if not qs:
                print(f"Create = {group_name}")
                
                qs = UserGroup.objects.create(name=group_name
                )
                
            for permission in custom_permission_list:
                name = permission.get('name')
                f_qs = CustomPermission.objects.filter(name__icontains=name).first()
                
                if not f_qs:
                    codename = name.lower().replace(" ", "_")   
                    f_qs = CustomPermission.objects.create(
                        name=name,
                        codename=codename,
                        )
                    
                f_qs = CustomPermission.objects.filter(name__icontains=name).last()
                
                qs.custom_permission.add(f_qs)
                
        return ResponseWrapper(msg='Success', status=200)
    
    
    def bulk_ecommerce_order_create(self, request, *args, **kwargs):
        company_qs = Company.objects.filter().last()
        
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        count = 5972
        total_rows = 15327  # Assuming the total number of rows is 123, adjust if needed
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=5973, values_only=True), start=1):
            rest_of = total_rows - count
            
            id = row[0]
            order_date = row[1]
            invoice_no = row[2]
            customer_name = row[3]
            mobile_number = row[4]
            email = row[5]
            address = row[6]
            courier_name = row[7]
            store_name = row[8]
            order_status = row[9]
            order_type = row[10]
            payment_type = row[11]
            note = row[12]
            payment_status = row[13]
            delivery_type = row[14]
            delivery_method = row[15]
            product_name = row[16] #Ok
            # brand = row[17]
            # seller = row[18]
            msp = row[19]
            # mrp = row[19]
            product_code = row[21]
            sku_list = row[22] #dd
            quantity = row[23] 
            mrp = row[24]
            # promo_discount_amount = row[25]
            # after_promo_selling_price = row[26]
            # promo_code = row[27]
            # commission_amount = row[28]
            total_product_price = row[25]
            total_discount_amount = row[26]
            total_net_payable_amount = row[27]
            total_gsheba_amount = row[28]
            total_delivery_charge = row[29]
            total_tax_amount = row[30]
            total_promo_discount = row[31]
            total_advance_amount = row[32]
            total_due_amount = row[33]
            total_payable_amount = row[34]  
            
            print(f"SI = {count}, Total = {total_rows}, Rest Of = {rest_of}, Invoice No  = {invoice_no}")
            count += 1
            
            if not payment_status:
                payment_status = "UNPAID"
            
            order_qs = Order.objects.filter(invoice_no = invoice_no).last()
            
            store_name_qs = None
            area_qs = None
            pickup_shop_qs = None
            user_qs = None
            delivery_method_qs = None
            customer_qs = None
            
            area_name = "-"
            district_name = "-"
            division_name = "-"
            country_name = "Bangladesh"
            # country_name = "-"
            
            # Order Type Start
            
            if order_type == "E-Commerce Sell":
                order_type = "ECOMMERCE_SELL"
                
            else:
                order_type = "RETAIL_ECOMMERCE_SELL"
            
            # Order Type End
            
            # Order Status Start
            
            order_status = order_status.replace(" ", "_").upper()
            
            # Order Status End
            
            # Order Delivery Method Start
            
            if delivery_method == "store_sale":
                delivery_method = "STORE_SELL"
                
            elif delivery_method == "store_pickup":
                delivery_method = "SHOP_PICKUP"
                
            elif delivery_method == "home_delivery":
                delivery_method = "HOME_DELIVERY"
                
            delivery_method_qs = DeliveryMethod.objects.filter(delivery_type = delivery_method).last()
                
            # else:
            #     delivery_method = "ORDER_CONFIRMED"
            
            # Order Delivery Method End
            
            approved_status = "APPROVED"
            
        
            if payment_status == "PAID":
                payment_status = "PAID"
                
            elif payment_status == "UNPAID":
                payment_status = "UNPAID"
            
            if payment_type == "Cash on Delivery":
                payment_type = "CASH_ON_DELIVERY"
                
            elif payment_type == "Online Payment":
                payment_type = "ONLINE_PAYMENT"
            
                
            if delivery_method == "SHOP_PICKUP":
                store_name_qs = OfficeLocation.objects.filter(name = store_name).last()
                
                if store_name_qs:
                    try:
                        area_qs = store_name_qs.area
                    except:
                        area_qs = Area.objects.filter(name__icontains = "Dhaka").last()
                
            # if order_type == "POINT_OF_SELL":
            #     pickup_shop_qs = store_name_qs
                
            address_parts = address.split(',')

            # Extract and clean each component
            division_name = address_parts[0].split('-')[1].strip()
            district_name = address_parts[1].split('-')[1].strip()
            area_name = address_parts[2].split('-')[1].strip()
            remaining_address = ','.join(address_parts[3:]).strip()
            
            customer_qs = UserInformation.objects.filter(
                Q(user__phone = mobile_number)
                |Q(user__email = email)
            ).last()
            
            # print('ggggggggggg', customer_qs)
            
            if not customer_name:
                customer_name = "Walk in Way Customer"
            
            if not customer_qs:
                print(f"Line 3996")
                
                user_qs = UserAccount.objects.filter(
                    Q(phone = mobile_number)
                    |Q(email = email)
                ).last()
                
                if not user_qs:  
                    user_qs = UserAccount.objects.create(
                        first_name = customer_name,
                        last_name = customer_name,
                        email = email,
                        phone = mobile_number
                    )
                    
                if user_qs:
                    user_qs.first_name = customer_name
                    user_qs.last_name = customer_name
                    user_qs.email = email
                    user_qs.phone = mobile_number
                    user_qs.save()
                    
                user_type_qs = UserType.objects.filter(
                    name__icontains = "Customer"
                ).last()
                
                customer_qs = UserInformation.objects.filter(user = user_qs).last()
                
                if not customer_qs:
                    
                    slug = unique_slug_generator(name= f"walk-in-way-customer-{count}")
                    
                    customer_qs = UserInformation.objects.create(
                        user = user_qs,
                        name = customer_name,
                        slug = slug,
                        address = address,
                        image = settings.NOT_FOUND_IMAGE,
                        user_type = user_type_qs,
                        created_by = request.user
                    )
                print(f"Line 4037")
            
            elif customer_qs:
                # customer_qs = UserInformation.objects.filter(user = user_qs).last()
                print(f"Line 4041")
                
                try:
                    if not customer_name:
                        print(f"Line 4045")
                        customer_name = "Walk in Way Customer"
                    
                    if customer_qs:
                        print(f"Line 4049")
                        customer_qs.name = customer_name
                        customer_qs.save()
                except:
                    pass
                
            customer_qs = UserInformation.objects.filter(name = customer_name).last()
            
            # customer_qs = UserInformation.objects.filter(user = user_qs).last()
            
            print('Customer Name =', customer_qs.name)
            
            if order_qs:
                # print(f"{invoice_no} Order Update")
                qs = Order.objects.filter(invoice_no = invoice_no)
                
                qs = qs.update(
                    invoice_no = invoice_no,
                    order_type = order_type,
                    # status = order_status,
                    approved_status = approved_status,
                    delivery_method = delivery_method_qs,
                    payment_status = payment_status,
                    payment_type = payment_type,
                    area = area_qs,
                    shop = store_name_qs,
                    pickup_shop = pickup_shop_qs,
                    customer = customer_qs,
                    district_name = district_name,
                    division_name = division_name,
                    country_name = country_name,
                    address = address,
                    remarks = note,
                    # promo_code = promo_code,
                    total_product_price = total_product_price,
                    total_discount_amount = total_discount_amount,
                    total_net_payable_amount = total_net_payable_amount,
                    total_gsheba_amount = total_gsheba_amount,
                    total_tax_amount = total_tax_amount,
                    total_promo_discount = total_promo_discount,
                    total_delivery_charge = total_delivery_charge,
                    total_payable_amount = total_payable_amount,
                    total_advance_amount = total_advance_amount,
                    total_paid_amount = total_advance_amount,
                    total_due_amount = total_due_amount,
                    order_date = order_date,
                    created_at = order_date,
                    created_by = request.user
                )
                
            
            if not order_qs:
                # print(f"{invoice_no} Order Create")
                
                order_qs = Order.objects.create(
                    invoice_no = invoice_no,
                    order_type = order_type,
                    status = order_status,
                    approved_status = approved_status,
                    delivery_method = delivery_method_qs,
                    payment_status = payment_status,
                    payment_type = payment_type,
                    area = area_qs,
                    shop = store_name_qs,
                    customer = customer_qs,
                    pickup_shop = pickup_shop_qs,
                    district_name = district_name,
                    division_name = division_name,
                    country_name = country_name,
                    address = address,
                    remarks = note,
                    # promo_code = promo_code,
                    total_product_price = total_product_price,
                    total_discount_amount = total_discount_amount,
                    total_net_payable_amount = total_net_payable_amount,
                    total_gsheba_amount = total_gsheba_amount,
                    total_tax_amount = total_tax_amount,
                    total_promo_discount = total_promo_discount,
                    total_delivery_charge = total_delivery_charge,
                    total_payable_amount = total_payable_amount,
                    total_advance_amount = total_advance_amount,
                    total_paid_amount = total_advance_amount,
                    total_due_amount = total_due_amount,
                    # total_promo_discount = promo_discount_amount,
                    order_date = order_date,
                    created_at = order_date,
                    created_by = request.user,
                )
                
            order_qs = Order.objects.filter(invoice_no = invoice_no).last()
            delivery_service_qs = None
            
            try:
                if courier_name == "GProjukti.com Delivery":
                    if store_name:
                        delivery_service_qs = CourierService.objects.filter(
                            name__icontains = store_name
                            ).last()
                else:
                    delivery_service_qs = CourierService.objects.filter(
                        name__icontains = courier_name
                        ).last()
                    
            except:
                pass
                
        
            courier_qs = Courier.objects.filter(
                order = order_qs
            ).last()
            
            if courier_qs:
                courier_qs.order = order_qs
                courier_qs.courier_service = delivery_service_qs
                courier_qs.save()
                
            else:
                slug = unique_slug_generator(name = f"{courier_name}-{order_qs.invoice_no}")
                courier_qs = Courier.objects.create(
                    order = order_qs,
                    slug = slug,
                    courier_service = delivery_service_qs,
                    created_by = request.user,
                    created_at = order_qs.order_date
                    
                )
                
            
            # print(f'Customer Name = {order_qs.customer},{customer_qs}')
            
            customer_address_log_qs = CustomerAddressInfoLog.objects.filter(
                order = order_qs
            ).last()
            
            address_type = "HOME"
            
            if mobile_number:
                if customer_address_log_qs:
                    customer_address_log_qs.order = order_qs
                    customer_address_log_qs.address_type = address_type
                    customer_address_log_qs.name = customer_name
                    customer_address_log_qs.phone = customer_qs.user.phone
                    customer_address_log_qs.email = customer_qs.user.email
                    customer_address_log_qs.address = address
                    customer_address_log_qs.area_name = area_name
                    customer_address_log_qs.district_name = district_name
                    customer_address_log_qs.division_name = division_name
                    customer_address_log_qs.country_name = country_name
                    
                    customer_address_log_qs.save()
                    
                if not customer_address_log_qs:
                    slug = unique_slug_generator(name = order_qs.customer.name)
                    
                    customer_address_log_qs = CustomerAddressInfoLog.objects.create(
                        order = order_qs,
                        slug = slug,
                        address_type = address_type,
                        name = customer_qs.name,
                        phone = customer_qs.user.phone,
                        email = customer_qs.user.email,
                        address = address,
                        area_name = area_name,
                        district_name = district_name,
                        division_name = division_name,
                        country_name = country_name,
                        created_by = request.user,
                    )
            
            # Order Payment Log
            
            customer_payment_log_qs = OrderPaymentLog.objects.filter(
                order = order_qs
            ).last()
            
            print("Ggggggggggggggg", payment_status, payment_type)
            
            # payment_status = "RECEIVED"
            
            # if payment_status == "UNPAID":
            #     if customer_payment_log_qs:
            #         customer_payment_log_qs.delete()
            
            if payment_status == "PAID":
                # print('ttttttttttttt')
                
                if payment_type == "CASH_ON_DELIVERY":
                    # print('uuuuuuuuuu')
                    order_payment_qs = PaymentType.objects.filter(
                        name__icontains ="Cash On Delivery"
                    ).last()
                    
                    remarks = f"Payment Received Date Is Not Found, So That Now Add Order Date as Payment Received Date By MIS Department"
                    
                    customer_payment_log_qs = OrderPaymentLog.objects.filter(
                        order = order_qs
                    ).last()
                    
                    # print('dfdsf')
                    
                    
                    if customer_payment_log_qs:
                        
                        # print('fffffffffffffff')
                        
                        customer_payment_log_qs.order_payment = order_payment_qs
                        customer_payment_log_qs.remarks = remarks
                        customer_payment_log_qs.created_at = order_qs.order_date
                        customer_payment_log_qs.save()
                        
                        print(f"Payment Type = {customer_payment_log_qs.order_payment}, order_payment_qs = {order_payment_qs}")
                        
                        # print(f"Payment Type = {customer_payment_log_qs.order_payment.name} {customer_payment_log_qs.remarks} and ID  = {customer_payment_log_qs.id}")
                        
                        # customer_payment_log_qs.order = order_qs
                        # customer_payment_log_qs.order_payment = order_payment_qs
                        # customer_payment_log_qs.received_amount = total_payable_amount
                        # customer_payment_log_qs.created_at = order_qs.order_date
                        # customer_payment_log_qs.remarks = remarks
                        # customer_address_log_qs.save()
                        
                        # print(f"Update......., {customer_payment_log_qs.remarks}")
                        
                    else:
                        slug = unique_slug_generator(name = order_qs.customer.name)
                        
                        
                        customer_payment_log_qs = OrderPaymentLog.objects.create(
                            order = order_qs,
                            slug = slug,
                            order_payment = order_payment_qs,
                            received_amount = total_payable_amount, 
                            created_by = request.user,
                            remarks = remarks,
                            created_at = order_qs.order_date
                        )
                        # print(f"Create......., {customer_payment_log_qs.remarks}")
                
            try:
                customer_payment_log_qs = OrderPaymentLog.objects.filter(
                    order = order_qs
                ).last()
                
                print(f"Payment Type = {customer_payment_log_qs.order_payment.name} {customer_payment_log_qs.remarks} and ID  = {customer_payment_log_qs.id}")
            except:
                pass
            
            product_qs = Product.objects.filter(
                name__icontains = product_name
            ).last()
            
            # print(f"product_qs ={product_code}, {product_name}")
            
            # print(f"product_qs ={product_code}, {product_name} {product_qs}")
            
            if not msp:
                msp = mrp
            
            if product_qs:
                order_item_qs = OrderItem.objects.filter(
                    order = order_qs, product =  product_qs
                ).last()

            # try:
                if order_item_qs:
                    order_item_qs = OrderItem.objects.filter(
                        order = order_qs
                    )
                    qs = order_item_qs.update(
                        order = order_qs,
                        status = order_qs.status,
                        quantity = quantity,
                        product = product_qs,
                        unit_msp_price = msp,
                        unit_mrp_price = mrp,
                        selling_price = mrp,
                        total_product_price = (total_product_price*quantity),
                        total_tax_amount = total_tax_amount,
                        commission_amount = 0.0,
                        barcode_number = "-",
                        remarks = str(set(sku_list)),
                        promo_code = '-',
                        
                        created_by = request.user,
                        created_at = order_date
                    )
                    
                if not order_item_qs:
                    
                    order_item_qs = OrderItem.objects.create(
                        order = order_qs,
                        status = order_qs.status,
                        quantity = quantity,
                        product = product_qs,
                        unit_msp_price = msp,
                        unit_mrp_price = mrp,
                        selling_price = mrp,
                        total_product_price = (total_product_price*quantity),
                        total_tax_amount = total_tax_amount,
                        commission_amount = 0.0,
                        barcode_number = "-",
                        remarks = str(set(sku_list)),
                        promo_code = '',
                        
                        created_by = request.user,
                        created_at = order_date
                    )
            # except:
            #     pass

        # print('Order Create Done')
        
                
                
        return ResponseWrapper(msg='Success', status=200)
    
    
    # This Function is For From 14 July To 28 July
    
    def bulk_shop_sell_create(self, request, *args, **kwargs):
        company_qs = Company.objects.filter().last()
        
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        count = 1
        total_rows = 8  
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            rest_of = total_rows - count
            
            # print(f"SI = {count}, Total = {total_rows}, Rest Of = {rest_of}")
            count += 1
            
            order_date = row[0]
            invoice_no = row[1]
            barcode = row[4]
            
            order_qs = Order.objects.filter(invoice_no = invoice_no).last()
            
            if not order_qs:
                print(f"{invoice_no} Invoice No is Not Found")
                
            else:
                print(f"....***.... {invoice_no} Invoice No is Found & Barcode Number = {barcode} & Order Barcode is  = {order_qs.order_items.last().barcode_number}, Order Date = {order_qs.order_date} and Date = {order_date} ....***....")
                order_qs.order_date = order_date
                order_qs.save()
                
                product_stock_log_qs = ProductStockLog.objects.filter(
                    product_stock__barcode = order_qs.order_items.last().barcode_number, current_status = "SOLD"
                ).last()
                
                product_stock_log_qs.created_at = order_qs.order_date
                product_stock_log_qs.save()
                
    
    def bulk_stock_transfer_update(self, request, *args, **kwargs):
        company_qs = Company.objects.filter().last()
        
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        count = 1
        total_rows = 8  
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            rest_of = total_rows - count
            
            # print(f"SI = {count}, Total = {total_rows}, Rest Of = {rest_of}")
            count += 1
            
            order_date = row[0]
            requisition_no = row[1]
            barcode = row[2]
            
            if not order_date:
                break
            
            qs = ProductStockTransfer.objects.filter(requisition_no = requisition_no).first()
            
            if not qs:
                print(f"{requisition_no} Requisition No is Not Found")
                
            else:
                print(f"....***.... {requisition_no} Invoice No is Found & Barcode = {barcode} ***....")
                qs.created_at = order_date
                qs.save()
                
                product_stock_log_qs = ProductStockLog.objects.filter(
                    product_stock__barcode = barcode, current_status = "IN_TRANSFER"
                ).first()
                
                if product_stock_log_qs:
                    print('gggggggg')
                    product_stock_log_qs.created_at = order_date
                    product_stock_log_qs.save()
                
                
            
        return ResponseWrapper(msg='Success', status=200)
    
    def bulk_barcode_status_update(self, request, *args, **kwargs):
        company_qs = Company.objects.filter().last()
        
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        count = 1
        total_rows = 7189  
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            rest_of = total_rows - count
            
            print(f"SI = {count}, Total = {total_rows}, Rest Of = {rest_of}")
            count += 1
            
            si = row[0]
            barcode = row[1]
            status = row[2]
            
            # if status == "active":
            #     pass
            
            msg_2 = '-'
            msg = '-'
            
            stock_qs = ProductStock.objects.filter(barcode = barcode).last()
            
            current_status = status.upper()
            
            if not stock_qs:
                print(f"{barcode} Barcode is Not Found")
                
            else:
                if stock_qs.stock_location:
                    print(f"....***.... {barcode} Barcode is Found & and Location {stock_qs.stock_location.name} and Current Status {stock_qs.get_status_display()}, V1 Status = {current_status} ....***....")
                else:
                    print(f"....###.... {barcode} Barcode is Found & and Location {stock_qs.stock_location} and Current Status {stock_qs.get_status_display()}, V1 Status = {current_status} ....###...")
                    
                # if not current_status == "ACTIVE":
                #     print(f'Updating To {current_status} and Location {stock_qs.stock_location} and Current Status = {stock_qs.get_status_display()}')
                    
                #     stock_qs.status = current_status
                #     stock_qs.save()
                    
                    
            # print(f" {msg_q} and {msg_2}")
            # print(f" {msg_q} and {msg} ... {msg_2}")
            
        return ResponseWrapper(msg='Success', status=200)
    
    
    
    @log_activity
    def bulk_day_end_create(self, request, *args, **kwargs):
        csv_file = request.FILES['file']
        decoded_file = csv_file.read().decode('utf-8').splitlines()
        reader = csv.reader(decoded_file)
        
        # Skip the header row
        headers = next(reader)
        
        count = 1
        total = 10538
        
        for row_number, row in enumerate(reader, start=1):
            si = row[0]
            day_end_time = row[1]
            total_shop_sale = convert_null_to_zero(row[2])
            total_ecommerce_collection = convert_null_to_zero(row[3])
            total_prothoma_collection = convert_null_to_zero(row[4])
            petty_cash = convert_null_to_zero(row[5])
            opening_amount = convert_null_to_zero(row[6])
            total_bank_deposit = convert_null_to_zero(row[7])
            total_expenses = convert_null_to_zero(row[8])
            balance = convert_null_to_zero(row[9])
            note = row[10]
            store_id = row[11]
            b2b = convert_null_to_zero(row[12])
            mfs = row[13]
            gsheba_claim_qty = convert_null_to_zero(row[14])
            refund_amount = convert_null_to_zero(row[15])
            warranty_claim_qty = convert_null_to_zero(row[16])
            is_message_sent = row[17]
            ssl_response_body = row[18]
            id_2 = row[19]
            created_at = row[20]
            updated_at = row[21]
            name = row[22]
            type_ = row[23]  # 'type' is a reserved keyword in Python, so using 'type_'
            slug = row[24]
                
                
            store_qs = OfficeLocation.objects.filter(slug=slug).last()
            
            day_end_time_str = day_end_time
            day_end_time = datetime.fromisoformat(day_end_time_str)
            
            day_end_qs = ShopDayEnd.objects.filter(
                shop=store_qs, day_end_date__date=day_end_time.date()
            ).last()
            
            panel_partnership = [
                {
                    "name": "Nazat", 
                    "total_sell_amount": total_prothoma_collection,
                    "total_gsheba_amount": 0
                }
            ]
            
            currency_collection = [
                {"currency": "1000", "quantity": 0}, {"currency": "500", "quantity": 0}, 
                {"currency": "200", "quantity": 0}, {"currency": "100", "quantity": 0}, 
                {"currency": "50", "quantity": 0}, {"currency": "20", "quantity": 0}, 
                {"currency": "10", "quantity": 0}, {"currency": "5", "quantity": 0}, 
                {"currency": "2", "quantity": 0}, {"currency": "1", "quantity": 0}
            ]
            
            mfs_collection = MFS_COLLECTION
            qs = day_end_qs
            
            if day_end_qs:
                if not day_end_qs.slug:
                    slug = unique_slug_generator(name=name)
                elif day_end_qs.slug:
                    slug = day_end_qs.slug

                # Update ShopDayEnd objects
                day_end_qs = ShopDayEnd.objects.filter(
                    shop=store_qs, day_end_date__date=day_end_time.date()
                )
                
                day_end_qs = day_end_qs.update(
                    shop=store_qs,
                    slug=slug,
                    retail_sell_amount=total_shop_sale,
                    panel_partnership=panel_partnership,
                    e_retail_sell_amount=0.0,
                    ecommerce_collection_amount=total_ecommerce_collection, 
                    refund_amount=refund_amount,
                    warranty_claim_quantity=warranty_claim_qty,
                    gsheba_claim_quantity=gsheba_claim_qty,
                    total_b2b_sell_amount=b2b, 
                    currency_collection=currency_collection,
                    total_bank_deposit_amount=total_bank_deposit,
                    total_expense_amount=total_expenses,
                    petty_cash_amount=petty_cash,
                    opening_balance_amount=opening_amount,
                    current_balance_amount=balance,
                    remarks=note,
                    day_end_date=day_end_time.date(),
                    created_at=day_end_time.date(),
                    created_by=request.user,
                    mfs_collection=mfs_collection
                )
                
            if not day_end_qs:
                slug = unique_slug_generator(name=name)
                
                day_end_qs = ShopDayEnd.objects.create(
                    shop=store_qs,
                    slug=slug,
                    retail_sell_amount=total_shop_sale,
                    panel_partnership=panel_partnership,
                    e_retail_sell_amount=0.0,
                    ecommerce_collection_amount=total_ecommerce_collection, 
                    refund_amount=refund_amount,
                    warranty_claim_quantity=warranty_claim_qty,
                    gsheba_claim_quantity=gsheba_claim_qty,
                    total_b2b_sell_amount=b2b, 
                    currency_collection=currency_collection,
                    total_bank_deposit_amount=total_bank_deposit,
                    total_expense_amount=total_expenses,
                    petty_cash_amount=petty_cash,
                    opening_balance_amount=opening_amount,
                    current_balance_amount=balance,
                    remarks=note,
                    day_end_date=day_end_time.date(),
                    created_at=day_end_time.date(),
                    created_by=request.user,
                    mfs_collection=mfs_collection
                )
                
            day_end_qs = ShopDayEnd.objects.filter(
                shop=store_qs, day_end_date__date=day_end_time.date()
            ).last()
            
            day_end_qs.day_end_date = day_end_time.date()
            day_end_qs.created_at = day_end_time.date()
            day_end_qs.save()
            
            rest_of = total - count
                
            print(f"ID = {count}, Rest Of = {rest_of}, Name = {name}, Day End = {day_end_time.date()}, Save Data is = {day_end_qs.day_end_date}")
            
            count +=1
            
        return ResponseWrapper(msg='Success', status=200)
    
    @log_activity
    def bulk_corporate_sell_create(self, request, *args, **kwargs):
        csv_file = request.FILES['file']
        decoded_file = csv_file.read().decode('utf-8').splitlines()
        reader = csv.reader(decoded_file)
        
        # Skip the header row
        headers = next(reader)
        
        count = 1
        total = 5517
        
        for row_number, row in enumerate(reader, start=1):
            si = row[0]
            order_status = row[1]
            product_name = row[2]
            barcode = row[3]
            msp = row[4]
            mrp = row[5]
            corporate_price = row[6]
            approved_date = row[7]
            shop_slug = row[9]
            
            if not si:
                break
            
            if approved_date == "NULL":
                approved_date = timezone.now() - timedelta(days=30)
            else:
                order_date = approved_date
            
            if corporate_price == "NULL":
                corporate_price = 0
            else:
                corporate_price = corporate_price
            
            customer_first_name = 'Corporate'
            customer_last_name = 'Customer'
            
            
            invoice_no = None
            
            order_qs = Order.objects.filter(order_items__barcode_number = barcode).last()
            
            if not order_qs:
                last_order_qs = Order.objects.all().order_by('-created_at').last()
        
                last_invoice_no = last_order_qs.invoice_no if last_order_qs else 'ONL000000001'
                invoice_no = generate_invoice_no(last_invoice_no)
                
                while not invoice_no:
                    invoice_no = generate_invoice_no(last_invoice_no)
                    qs = Order.objects.filter(invoice_no=invoice_no).last()
                    if qs:
                        product_code = None
                
            else:
                invoice_no = order_qs.invoice_no
            
            store_name_qs = None
            area_qs = None
            pickup_shop_qs = None
            user_qs = None
            delivery_method_qs = None
            customer_qs = None
            
            area_name = "-"
            district_name = "-"
            division_name = "-"
            country_name = "Bangladesh"
            
            
            # Order Type Start
            
            order_type = "CORPORATE_SELL"
            
            # Order Type End
            
            # Order Status Start
            
            if order_status == "approved":
                order_status = "DELIVERED"
            else:
                order_status = "ORDER_RECEIVED"
                
            # else:
            #     try:
            #         order_status = order_status.replace(" ", "_").upper()
            #     except:
            #         order_status = "DELIVERED"
                    
                # order_status = "ORDER_CONFIRMED"
            
            # Order Status End
            
            # Order Delivery Method Start
            
            # delivery_method == "store_pickup":
            
            delivery_method = "SHOP_PICKUP"
                
            delivery_method_qs = DeliveryMethod.objects.filter(delivery_type = delivery_method).last()
                
            # else:
            #     delivery_method = "ORDER_CONFIRMED"
            
            # Order Delivery Method End
            
            if order_status == "DELIVERED":
                approved_status = "APPROVED"
            else:
                approved_status = "INITIALIZED"
                
            if approved_status == "APPROVED":
                order_date = approved_date
            
        
            if order_status == "DELIVERED":
                payment_status = "PAID"
                
            else:
                payment_status = "UNPAID"
            
            # if payment_type == "Cash on Delivery":
            
            payment_type = "CASH_ON_DELIVERY"
                
            # elif payment_type == "Online Payment":
            #     payment_type = "ONLINE_PAYMENT"
            
                
            store_name_qs = OfficeLocation.objects.filter(slug = shop_slug).last()
            
            if store_name_qs:
                try:
                    area_qs = store_name_qs.area
                except:
                    area_qs = Area.objects.filter(name__icontains = "Dhaka").last()
                
            if order_type == "POINT_OF_SELL":
                pickup_shop_qs = store_name_qs
                
            if area_qs:
                area_name = area_qs.name
                district_name = area_qs.district.name
                division_name = area_qs.district.division.name
                country_name = "Bangladesh"
                
            address = f"Area = {area_name}, District: ={district_name}, Division = {division_name}, Country = {country_name}"
            
            mobile_number = "corporate_sell@gmail.com"
            email = "corporate_sell@gmail.com"
            
            customer_qs = UserInformation.objects.filter(
                Q(user__phone = mobile_number)
                |Q(user__email = email)
            ).last()
            
            # print('ggggggggggg', customer_qs)
            
            # if not customer_name:
            customer_name = "Corporate Customer"
            
            if not customer_qs:
                # print(f"Line 3996")
                
                user_qs = UserAccount.objects.filter(
                    Q(phone = mobile_number)
                    |Q(email = email)
                ).last()
                
                if not user_qs:  
                    user_qs = UserAccount.objects.create(
                        first_name = customer_name,
                        last_name = customer_name,
                        email = email,
                        phone = mobile_number
                    )
                    
                if user_qs:
                    user_qs.first_name = customer_name
                    user_qs.last_name = customer_name
                    user_qs.email = email
                    user_qs.phone = mobile_number
                    user_qs.save()
                    
                user_type_qs = UserType.objects.filter(
                    name__icontains = "Customer"
                ).last()
                
                customer_qs = UserInformation.objects.filter(user = user_qs).last()
                
                if not customer_qs:
                    
                    slug = unique_slug_generator(name= f"walk-in-way-customer-{count}")
                    
                    customer_qs = UserInformation.objects.create(
                        user = user_qs,
                        name = customer_name,
                        slug = slug,
                        address = address,
                        image = settings.NOT_FOUND_IMAGE,
                        user_type = user_type_qs,
                        created_by = request.user
                    )
                # print(f"Line 4037")
            
            elif customer_qs:
                # customer_qs = UserInformation.objects.filter(user = user_qs).last()
                # print(f"Line 4041")
                
                try:
                    if not customer_name:
                        # print(f"Line 4045")
                        customer_name = "Corporate Customer"
                    
                    if customer_qs:
                        # print(f"Line 4049")
                        customer_qs.name = customer_name
                        customer_qs.save()
                except:
                    pass
                
            customer_qs = UserInformation.objects.filter(name__icontains = customer_name).last()
            
            if not customer_qs:
                customer_qs = UserInformation.objects.filter(name = "Corporate Customer").last()
                
            # customer_qs = UserInformation.objects.filter(user = user_qs).last()
            
            # print('Customer Name =', customer_qs.name)
            
            note ="This is Generated By MIS Department"
            
            promo_code= ''
            total_product_price = mrp
            total_discount_amount = 0.0
            total_net_payable_amount = corporate_price
            total_gsheba_amount = 0.0
            total_tax_amount = 0.0
            total_promo_discount = 0.0
            total_delivery_charge = 0.0
            total_gsheba_amount = 0.0
            total_payable_amount = corporate_price
            total_paid_amount = 0.0
            total_advance_amount = 0.0
            # order_date = approved_date
            total_due_amount = 0.0
            total_paid_amount = 0.0
            
            if order_qs:
                print(f"{invoice_no} Order Update")
                qs = Order.objects.filter(invoice_no = invoice_no)
                
                qs = qs.update(
                    invoice_no = invoice_no,
                    order_type = order_type,
                    status = order_status,
                    approved_status = approved_status,
                    delivery_method = delivery_method_qs,
                    payment_status = payment_status,
                    payment_type = payment_type,
                    area = area_qs,
                    shop = store_name_qs,
                    pickup_shop = pickup_shop_qs,
                    customer = customer_qs,
                    district_name = district_name,
                    division_name = division_name,
                    country_name = country_name,
                    address = address,
                    remarks = note,
                    promo_code = promo_code,
                    total_product_price = total_product_price,
                    total_discount_amount = total_discount_amount,
                    total_net_payable_amount = total_net_payable_amount,
                    total_gsheba_amount = total_gsheba_amount,
                    total_tax_amount = total_tax_amount,
                    total_promo_discount = total_promo_discount,
                    total_delivery_charge = total_delivery_charge,
                    total_payable_amount = total_payable_amount,
                    total_advance_amount = total_advance_amount,
                    total_paid_amount = total_paid_amount,
                    total_due_amount = total_due_amount,
                    order_date = order_date,
                    created_at = order_date,
                    created_by = request.user
                )
                
            
            if not order_qs:
                print(f"{invoice_no} Order Create")
                
                # last_order_qs = Order.objects.all().order_by('-created_at').last()
        
                # last_invoice_no = last_order_qs.invoice_no if last_order_qs else 'ONL000000001'
                # invoice_no = generate_invoice_no(last_invoice_no)
                
                customer_qs = UserInformation.objects.filter(name = "Corporate Customer").last()
                
                order_qs = Order.objects.create(
                    invoice_no = invoice_no,
                    order_type = order_type,
                    status = order_status,
                    approved_status = approved_status,
                    delivery_method = delivery_method_qs,
                    payment_status = payment_status,
                    payment_type = payment_type,
                    area = area_qs,
                    shop = store_name_qs,
                    customer = customer_qs,
                    pickup_shop = pickup_shop_qs,
                    district_name = district_name,
                    division_name = division_name,
                    country_name = country_name,
                    address = address,
                    remarks = note,
                    promo_code = promo_code,
                    total_product_price = total_product_price,
                    total_discount_amount = total_discount_amount,
                    total_net_payable_amount = total_net_payable_amount,
                    total_gsheba_amount = total_gsheba_amount,
                    total_tax_amount = total_tax_amount,
                    total_promo_discount = total_promo_discount,
                    total_delivery_charge = total_delivery_charge,
                    total_payable_amount = total_payable_amount,
                    total_advance_amount = total_advance_amount,
                    total_paid_amount = total_advance_amount,
                    total_due_amount = total_due_amount,
                    # total_promo_discount = promo_discount_amount,
                    order_date = order_date,
                    created_at = order_date,
                    created_by = request.user,
                )
                
                print(f'Order Invoice No ={invoice_no}')
                
            order_qs = Order.objects.filter(invoice_no = invoice_no).last()
            
            print(f'Order Name ={order_qs.invoice_no}')
            
            print(f'Customer Name = {order_qs}, {order_qs.customer},{customer_qs}')
            
            customer_address_log_qs = CustomerAddressInfoLog.objects.filter(
                order = order_qs
            ).last()
            
            address_type = "HOME"
            
            if customer_address_log_qs:
                customer_address_log_qs.order = order_qs
                customer_address_log_qs.address_type = address_type
                customer_address_log_qs.name = customer_qs.name
                customer_address_log_qs.phone = customer_qs.user.phone
                customer_address_log_qs.email = customer_qs.user.email
                customer_address_log_qs.address = address
                customer_address_log_qs.area_name = area_name
                customer_address_log_qs.district_name = district_name
                customer_address_log_qs.division_name = division_name
                customer_address_log_qs.country_name = country_name
                
                customer_address_log_qs.save()
                
            if not customer_address_log_qs:
                slug = unique_slug_generator(name = customer_name)
                
                customer_address_log_qs = CustomerAddressInfoLog.objects.create(
                    order = order_qs,
                    slug = slug,
                    address_type = address_type,
                    name = customer_qs.name,
                    phone = customer_qs.user.phone,
                    email = customer_qs.user.email,
                    address = address,
                    area_name = area_name,
                    district_name = district_name,
                    division_name = division_name,
                    country_name = country_name,
                    created_by = request.user,
                )
            
            
            # Order Payment Log
            
            customer_payment_log_qs = OrderPaymentLog.objects.filter(
                order = order_qs
            ).last()
            
            if approved_status == "APPROVED":
                
                payment_status = "RECEIVED"
            
                order_payment_qs = PaymentType.objects.filter(
                    name__icontains ="Cash On Delivery"
                ).last()
            
                if customer_payment_log_qs:
                    customer_payment_log_qs.order = order_qs
                    customer_payment_log_qs.order_payment = order_payment_qs
                    customer_payment_log_qs.received_amount = total_payable_amount
                    customer_payment_log_qs.created_at = order_qs.order_date
                    customer_address_log_qs.save()
                    
                if not customer_payment_log_qs:
                    slug = unique_slug_generator(name = order_qs.customer.name)
                    
                    customer_payment_log_qs = OrderPaymentLog.objects.create(
                        order = order_qs,
                        slug = slug,
                        order_payment = order_payment_qs,
                        received_amount = total_payable_amount,
                        created_at = order_qs.order_date,
                        created_by = request.user,
                    )
                    
                order_item_qs = OrderItem.objects.filter(
                    order = order_qs
                ).last()
                
                product_qs = Product.objects.filter(
                    name = product_name
                ).last()
            
            # sku_list = list(set(sku_list))  # Ensure there are no duplicates
            # formatted_sku_list = []

            # for sku in sku_list:
            #     for char in sku:
            #         formatted_sku_list.append(f'[[\\"{char}\\"]]')

            # barcode = ', '.join(formatted_sku_list)
            
            # barcode = next((barcode for barcode in sku_list if barcode.startswith(product_code)))
            
            # barcode_list = ast.literal_eval(sku_list)
            
            # barcode = next((barcode for barcode in barcode_list if barcode.startswith(product_code)), "-")
            
            barcode = barcode
            
            # print(f"Barcode = {barcode}, product_code={product_code} sku  = {sku_list}")
            
            try:
                product_stock_qs = ProductStock.objects.filter(barcode = barcode).last()
                
                stock_location_qs = order_qs.shop
                # print(f"SHop = {order_qs.shop}")
                
                date = order_qs.order_date.date() + timedelta(days=1)
                
                remarks = f"This Barcode History is Generated By MIS Department at {date} and Order Invoice No is {invoice_no}, Also Sold From {stock_location_qs.name}"
                
                if product_stock_qs:
                    # print(f"{barcode} is Update ....")
                    
                    product_stock_qs.status = "SOLD"
                    product_stock_qs.stock_location = stock_location_qs
                    product_stock_qs.created_at = order_qs.order_date
                    product_stock_qs.created_by = order_qs.created_by
                    product_stock_qs.remarks = remarks
                    product_stock_qs.save()
                    
                    product_stock_log_qs = ProductStockLog.objects.filter(
                    product_stock = product_stock_qs,
                    current_status = "SOLD"
                    ).last()
                    
                    # print(f"ffff= {product_stock_log_qs}")
                    
                    if not product_stock_log_qs:
                        barcode_status_log(product_stock_qs = product_stock_qs, previous_status = 'ACTIVE', previous_status_display ='Active'  , current_status = "SOLD", remarks = product_stock_qs.remarks, is_active = True, request_user = order_qs.created_by,stock_in_date = order_qs.order_date)
                        
                    else:
                        product_stock_log_qs.remarks = remarks
                        product_stock_log_qs.created_at = order_qs.order_date
                        product_stock_log_qs.save()
                        
                        
                else:
                    # print(f"{barcode} is Creating ....")
                    product_price_info_qs = ProductPriceInfo.objects.filter(
                        product__slug = product_qs.slug,
                        product_price_type = "CORPORATE"
                    ).last()
                    
                    product_stock_qs = ProductStock.objects.create(
                        barcode = barcode,
                        product_price_info = product_price_info_qs,
                        status = "SOLD",
                        stock_location = stock_location_qs,
                        stock_in_date = order_qs.order_date,
                        created_at = order_qs.order_date,
                        created_by = order_qs.created_by,
                        remarks = remarks
                        )
                    
                    
                    product_stock_log_qs = ProductStockLog.objects.filter(
                    product_stock = product_stock_qs,
                    current_status = "SOLD"
                    ).last()
                    
                    # print(f"ffff= {product_stock_log_qs}")
                    
                    if not product_stock_log_qs:
                        barcode_status_log(product_stock_qs = product_stock_qs, previous_status = 'ACTIVE', previous_status_display ='Active'  , current_status = "SOLD", remarks = product_stock_qs.remarks, is_active = True, request_user = order_qs.created_by,stock_in_date = order_qs.order_date)
                        
                    else:
                        product_stock_log_qs.remarks = remarks
                        product_stock_log_qs.created_at = order_qs.order_date
                        product_stock_log_qs.save()
                
                
            except:
                pass
            

            try:
                quantity = 1
                selling_price = corporate_price
                commission_amount = 0.0
                
                if order_item_qs:
                    order_item_qs = OrderItem.objects.filter(
                        order = order_qs
                    )
                    qs = order_item_qs.update(
                        order = order_qs,
                        # status = order_qs.status,
                        quantity = quantity,
                        product = product_qs,
                        unit_msp_price = msp,
                        unit_mrp_price = mrp,
                        selling_price = selling_price,
                        total_product_price = (total_product_price*quantity),
                        total_tax_amount = total_tax_amount,
                        commission_amount = commission_amount,
                        barcode_number = barcode,
                        remarks = barcode,
                        promo_code = promo_code,
                        gsheba_amount = total_gsheba_amount,
                        
                        created_by = request.user,
                        created_at = order_date
                    )
                    
                if not order_item_qs:
                    
                    order_item_qs = OrderItem.objects.create(
                        order = order_qs,
                        status = order_qs.status,
                        quantity = quantity,
                        product = product_qs,
                        unit_msp_price = msp,
                        unit_mrp_price = mrp,
                        selling_price = selling_price,
                        total_product_price = (total_product_price*quantity),
                        total_tax_amount = total_tax_amount,
                        commission_amount = commission_amount,
                        barcode_number = barcode,
                        remarks = barcode,
                        promo_code = promo_code,
                        gsheba_amount = total_gsheba_amount,
                        
                        created_by = request.user,
                        created_at = order_date
                    )
                    
                if order_qs.status == "DELIVERED":
                    
                    order_item_qs = OrderItem.objects.filter(
                        order = order_qs
                    ).last()
                            
                    product_warranty_list = product_qs.product_warrantys.all().order_by('warranty_type')
                    
                    if product_warranty_list:
                        for product_warranty in product_warranty_list:
                            warranty_type = product_warranty.warranty_type
                            warranty_duration = product_warranty.warranty_duration
                            value = product_warranty.value
                            
                            start_date = order_qs.order_date
                            
                            if warranty_type == "1_GSHEBA_WARRANTY" and order_item_qs.gsheba_amount == 0.0:
                                pass
                            else:
                                order_item_warranty_qs = order_item_qs.order_item_warranty_logs.all().last()
                                
                                if order_item_warranty_qs:
                                    start_date = order_qs.order_date
                                    
                                if warranty_duration == 'DAY':
                                    end_date = start_date + timedelta(days=int(value))
                                elif warranty_duration == 'MONTH':
                                    end_date = start_date + timedelta(days=30 * int(value))  # Assuming 1 month = 30 days
                                elif warranty_duration == 'YEAR':
                                    end_date = start_date + timedelta(days=365 * int(value)) 
                                    
                                # print(f"Warranty Start Date = {start_date} & End Date = {end_date}")
                                
                                if order_item_warranty_qs:
                                    # print('AAAA')
                                    order_item_warranty_log_qs = order_item_qs.order_item_warranty_logs.filter(warranty_type = warranty_type)
                                    
                                    if not order_item_warranty_log_qs:
                                        # print('BBBB')
                                        order_item_warranty_qs = OrderItemWarrantyLog.objects.create(
                                            order_item = order_item_qs,
                                            warranty_type = warranty_type,
                                            warranty_duration = warranty_duration,
                                            value = value,
                                            start_date = start_date,
                                            end_date = end_date,
                                            created_by = order_qs.created_by
                                        )
                                    else:
                                        # print('DDDD')
                                        
                                        order_item_warranty_log_qs.update(
                                            order_item = order_item_qs,
                                            warranty_type = warranty_type,
                                            warranty_duration = warranty_duration,
                                            value = value,
                                            start_date = start_date,
                                            end_date = end_date,
                                            created_by = order_qs.created_by
                                        )
                                        
                                # elif not order_item_warranty_qs:
                                    # print('CCC')
                                    order_item_warranty_qs = OrderItemWarrantyLog.objects.create(
                                            order_item = order_item_qs,
                                            warranty_type = warranty_type,
                                            warranty_duration = warranty_duration,
                                            value = value,
                                            start_date = start_date,
                                            end_date = end_date,
                                            created_by = order_qs.created_by
                                        )
                                    
                                # print(f"After Warranty Start Date = {order_item_warranty_qs.start_date} & End Date = {order_item_warranty_qs.end_date}")
                        
            except:
                pass
            
            rest_of = total - count
            print(f"ID = {count}, Rest Of = {rest_of}, Date = {approved_date}")
            
            count +=1
            
        return ResponseWrapper(msg='Success', status=200)
    
    @log_activity
    def bulk_barcode_update(self, request, *args, **kwargs):
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        total = 10
        office_location_qs =OfficeLocation.objects.filter(slug = 'gprojukti-com-shibganj').last()
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            barcode = row[0]
            
            if not barcode:
                break 
            product_stock_qs = ProductStock.objects.filter(barcode = barcode).last()
            if product_stock_qs:
                print(f"Update")
                product_stock_qs.stock_location = office_location_qs
                product_stock_qs.status = "ACTIVE"
                product_stock_qs.save()
                
            else:
                print(f"Not Found")
            
            
            # rest_of = total - count
            # print(f"ID = {count}, Rest Of = {rest_of}")
            
            # count +=1
            
        return ResponseWrapper(msg='Success', status=200)
    
    def bulk_parental_product_create(self, request, *args, **kwargs):
        
        xlsx_file = request.FILES['file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active 
        
        total = 356
        count = 1
        
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            si = row[0]
            product_name = row[1]
            product_code = row[2]
            parent_product_name = row[3]
            
            print(f".......***.......SI = {count},  {product_code} is OK.......***.......")
            
            if not si:
                break 
            
            product_qs = Product.objects.filter(product_code = product_code).last()
            
            if product_qs:
                if product_qs.name == product_name:
                    
                    print(f".......***.......SI = {count},  {product_code} is OK.......***.......")
                    count +=1
                    
                    parent_product_qs = Product.objects.filter(name = parent_product_name).last()
                    
                    if not parent_product_qs:
                    
                        parent_product_code = None
            
                        while not parent_product_code:
                            parent_product_code = random.randint(10000, 99999)
                            parent_product_code_qs = Product.objects.filter(product_code=parent_product_code).last()
                            
                            if parent_product_code_qs:
                                parent_product_code = None
                                
                        # parent_product_qs = Product.objects.filter(product_code = parent_product_code).last()
                        
                        parent_product_slug = f"{parent_product_name}-{parent_product_code}"
                        
                        slug = unique_slug_generator_for_product_category(name = parent_product_slug)
                        
                        qs = Product.objects.create(
                            name=parent_product_name,
                            slug=slug,
                            status="PARENT",
                            translation=product_qs.translation,
                            specifications=product_qs.specifications,
                            meta=product_qs.meta,
                            short_description=product_qs.short_description,
                            description=product_qs.description,
                            minimum_stock_quantity=product_qs.minimum_stock_quantity,
                            is_featured=product_qs.is_featured,
                            is_top_sale=product_qs.is_top_sale,
                            is_upcoming=product_qs.is_upcoming,
                            is_new_arrival=product_qs.is_new_arrival,
                            is_on_the_go=product_qs.is_on_the_go,
                            is_out_of_stock=product_qs.is_out_of_stock,
                            is_cart_disabled=product_qs.is_cart_disabled,
                            is_gift_product=product_qs.is_gift_product,
                            is_special_day=product_qs.is_special_day,
                            show_on_landing_page=product_qs.show_on_landing_page,
                            is_active=product_qs.is_active,
                            is_commission_enable=product_qs.is_commission_enable,
                            remarks=product_qs.remarks,
                            rating=product_qs.rating,
                            integrity_guaranteed=product_qs.integrity_guaranteed,
                            product_code=parent_product_code,
                            banner_message=product_qs.banner_message,
                            # Related fields will have to be set after the object is created
                            company=product_qs.company,
                            brand=product_qs.brand,
                            supplier=product_qs.supplier,
                            seller=product_qs.seller,
                            selling_tax_category=product_qs.selling_tax_category,
                            buying_tax_category=product_qs.buying_tax_category,
                            images=product_qs.images,
                            sku=product_qs.sku,
                            video_link=product_qs.video_link,
                            serial_no=product_qs.serial_no,
                            meta_title=product_qs.meta_title,
                            meta_image=product_qs.meta_image,
                            meta_description=product_qs.meta_description,
                            og_title=product_qs.og_title,
                            og_image=product_qs.og_image,
                            og_url=product_qs.og_url,
                            og_description=product_qs.og_description,
                            canonical=product_qs.canonical,
                            banner_image=product_qs.banner_image,
                            created_by=product_qs.created_by,
                            updated_by=product_qs.updated_by,
                        )
                        
                        qs.category.set(product_qs.category.all())
                        qs.sub_category.set(product_qs.sub_category.all())

                        # Optionally, save the instance again if needed
                        qs.save()

                            
                    else:
                        parent_product_qs = Product.objects.filter(name = parent_product_name)
                        
                        if parent_product_qs:
                            parent_product_qs = parent_product_qs.update(
                            status="PARENT",
                            translation=product_qs.translation,
                            specifications=product_qs.specifications,
                            meta=product_qs.meta,
                            short_description=product_qs.short_description,
                            description=product_qs.description,
                            minimum_stock_quantity=product_qs.minimum_stock_quantity,
                            is_featured=product_qs.is_featured,
                            is_top_sale=product_qs.is_top_sale,
                            is_upcoming=product_qs.is_upcoming,
                            is_new_arrival=product_qs.is_new_arrival,
                            is_on_the_go=product_qs.is_on_the_go,
                            is_out_of_stock=product_qs.is_out_of_stock,
                            is_cart_disabled=product_qs.is_cart_disabled,
                            is_gift_product=product_qs.is_gift_product,
                            is_special_day=product_qs.is_special_day,
                            show_on_landing_page=product_qs.show_on_landing_page,
                            is_active=product_qs.is_active,
                            is_commission_enable=product_qs.is_commission_enable,
                            remarks=product_qs.remarks,
                            rating=product_qs.rating,
                            integrity_guaranteed=product_qs.integrity_guaranteed,
                            # product_code=parent_product_code,
                            banner_message=product_qs.banner_message,
                            # Related fields will have to be set after the object is created
                            company=product_qs.company,
                            brand=product_qs.brand,
                            supplier=product_qs.supplier,
                            seller=product_qs.seller,
                            selling_tax_category=product_qs.selling_tax_category,
                            buying_tax_category=product_qs.buying_tax_category,
                            images=product_qs.images,
                            sku=product_qs.sku,
                            video_link=product_qs.video_link,
                            serial_no=product_qs.serial_no,
                            meta_title=product_qs.meta_title,
                            meta_image=product_qs.meta_image,
                            meta_description=product_qs.meta_description,
                            og_title=product_qs.og_title,
                            og_image=product_qs.og_image,
                            og_url=product_qs.og_url,
                            og_description=product_qs.og_description,
                            canonical=product_qs.canonical,
                            banner_image=product_qs.banner_image,
                            created_by=product_qs.created_by,
                            updated_by=product_qs.updated_by,
                            )
                        
                    
            else:
                print(f"{product_code} Not OK")
                
            child_product_price_list_qs = product_qs.product_price_infos.all()
            
            if child_product_price_list_qs:
                parent_product_qs = Product.objects.filter(name = parent_product_name).last()
                
                for product_price in child_product_price_list_qs:
                    product = product_price.product
                    
                    discount = product_price.discount
                    promo_code = product_price.promo_code
                    product_price_type = product_price.product_price_type
                    buying_price = product_price.buying_price
                    gsheba_amount = product_price.gsheba_amount
                    msp = product_price.msp
                    mrp = product_price.mrp
                    advance_amount_type = product_price.advance_amount_type
                    advance_amount = product_price.advance_amount
                    is_active = product_price.is_active
                    created_at = product_price.created_at
                    updated_at = product_price.updated_at
                    created_by = product_price.created_by
                    updated_by = product_price.updated_by
                    
                    product_price_qs = ProductPriceInfo.objects.filter(
                        product = parent_product_qs,
                        product_price_type = product_price_type
                    )
                    
                    if not product_price_qs:
                        # Create new ProductPriceInfo entry if it doesn't exist
                        product_price_qs = ProductPriceInfo.objects.create(
                            product=parent_product_qs,
                            discount=discount,
                            promo_code=promo_code,
                            product_price_type=product_price_type,
                            buying_price=buying_price,
                            gsheba_amount=gsheba_amount,
                            msp=msp,
                            mrp=mrp,
                            advance_amount_type=advance_amount_type,
                            advance_amount=advance_amount,
                            is_active=is_active,
                            created_at=created_at,
                            updated_at=updated_at,
                            created_by=created_by,
                            updated_by=updated_by,
                        )
                    else:
                        # Update existing ProductPriceInfo entry if it already exists
                        product_price_qs.update(
                            discount=discount,
                            promo_code=promo_code,
                            buying_price=buying_price,
                            gsheba_amount=gsheba_amount,
                            msp=msp,
                            mrp=mrp,
                            advance_amount_type=advance_amount_type,
                            advance_amount=advance_amount,
                            is_active=is_active,
                            updated_at=updated_at,
                            updated_by=updated_by,
                        ) 
        
            product_qs.status = "CHILD"
            product_qs.product_parent = parent_product_qs
            
            product_qs.save()
        
        return ResponseWrapper(msg='Success', status=200)
    
    
    def bulk_service_order_create(self, request, invoice_no, *args, **kwargs):
        qs = Order.objects.filter(invoice_no = invoice_no).last()
        
        if not qs:
            return ResponseWrapper(error_msg='Invoice No is not Found', status=200)
        
        last_service_order_qs = ServicingOrder.objects.all().order_by('-created_at').last()
        
        if last_service_order_qs:
            last_invoice_no = last_service_order_qs.invoice_no
        else:
            last_invoice_no = 'SER000000001'
            
        invoice_no = generate_service_invoice_no(last_invoice_no)
        
        today = timezone.now()

        service_order_create_or_update(invoice_no = invoice_no,servicing_type = 'ORDER', order = qs, request_user = qs.created_by, status = 'WAREHOUSE_TO_SERVICE_POINT', order_date = today)
        
        return ResponseWrapper(msg='Success', status=200)
    
    